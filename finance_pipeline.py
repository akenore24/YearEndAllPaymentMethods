#!/usr/bin/env python3
"""
finance_pipeline.py (ONE DRY FILE) — UPDATED

Reads a bank-style expenses CSV and produces:
1) Excel DETAIL (.xlsx): cleaned + grouped rows with subtotal lines
2) Excel SUMMARY (.xlsx): Group | Txns | Total (sorted txns desc, A→Z ties)
3) PDF DETAIL (.pdf): grouped by group/family with tables
4) PDF SUMMARY (.pdf): Group | Txns | Total (sorted txns desc, A→Z ties)

Key improvements (raw-data fixes):
- Strips narration prefixes like "PURCHASE AUTHORIZED ON 09/08 ..." so merchants group correctly.
- Produces a "merchant core" so variations like "7-ELEVEN 21494 ..." and "PURCHASE ... 7-ELEVEN ..."
  land in the SAME group.

Folder setup:
- Put this script and expenses.csv in the same folder.

Run:
  python3 finance_pipeline.py
  python3 finance_pipeline.py --excel-summary
  python3 finance_pipeline.py --excel-detail --pdf-summary
"""

import csv
import re
import argparse
from pathlib import Path
from datetime import datetime

# Excel
from openpyxl import Workbook
from openpyxl.styles import Font

# PDF
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
)


# -----------------------------
# Defaults / Config
# -----------------------------
DEFAULT_INPUT_CSV = "expenses_raw_spacing_fixed.csv" #

DEFAULT_EXCEL_DETAIL_OUT = "expenses_clean_grouped.xlsx"
DEFAULT_EXCEL_SUMMARY_OUT = "expenses_family_summary.xlsx"

DEFAULT_PDF_DETAIL_OUT = "expenses_grouped_families_detail.pdf"
DEFAULT_PDF_SUMMARY_OUT = "expenses_family_summary_by_txns_AZ.pdf"

REMOVE_DESC_PREFIX = "ONLINE TRANSFER REF"

WF_CARD_PREFIX = "WELLS FARGO ACTIVE CASH VISA(R) CARD"
WF_CARD_ALIAS = "WFACV"

DATE_FORMATS = (
    "%m/%d/%Y",
    "%m/%d/%y",
    "%Y-%m-%d",
    "%m-%d-%Y",
    "%m-%d-%y",
)

BOLD = Font(bold=True)


# -----------------------------
# Core Normalizers / Parsers
# -----------------------------
def normalize_spaces(text: str) -> str:
    """Trim and collapse internal whitespace to a single space."""
    return " ".join((text or "").split()).strip()


def normalize_payment_method(value: str) -> str:
    """
    Normalize WF Active Cash Visa payment method:
    'WELLS FARGO ACTIVE CASH VISA(R) CARD ...4321' -> 'WFACV...4321'
    """
    value = (value or "").strip()
    if value.upper().startswith(WF_CARD_PREFIX):
        suffix = value[len(WF_CARD_PREFIX):].strip()
        return f"{WF_CARD_ALIAS}{suffix}"
    return value


def parse_date(value: str):
    """Parse common bank/Excel date formats. Returns datetime or None."""
    s = ("" if value is None else str(value)).strip()
    if not s:
        return None

    # strip time parts if present
    s = s.split()[0]
    if "T" in s:
        s = s.split("T")[0]

    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def parse_amount(value) -> float:
    """Convert Amount cell to float safely. Handles $ , and (negative) style."""
    if value is None:
        return 0.0

    s = str(value).strip()
    if not s:
        return 0.0

    negative = False
    if s.startswith("(") and s.endswith(")"):
        negative = True
        s = s[1:-1].strip()

    s = s.replace("$", "").replace(",", "")
    try:
        n = float(s)
        return -n if negative else n
    except ValueError:
        return 0.0


def fmt_money(n: float) -> str:
    """Format float as currency string for PDF."""
    return f"${n:,.2f}"


# -----------------------------
# Description cleaning (raw narration fix)
# -----------------------------
def clean_description(raw: str) -> str:
    """
    Normalize the Description and strip common bank narration prefixes so the real
    merchant text remains. Works for multiple patterns seen in bank exports.

    Examples:
      "PURCHASE AUTHORIZED ON 09/08 7-ELEVEN Aurora CO"
        -> "7-ELEVEN Aurora CO"
      "PURCHASE AUTHORIZED ON 03/28 COSTCO GAS #1652 DENVER"
        -> "COSTCO GAS #1652 DENVER"
    """
    d = normalize_spaces(raw)
    if not d:
        return ""

    # PURCHASE AUTHORIZED ON MM/DD <merchant...>
    m = re.match(r"^PURCHASE\s+AUTHORIZED\s+ON\s+\d{2}/\d{2}\s+(.*)$", d, flags=re.IGNORECASE)
    if m:
        return m.group(1).strip()

    # ATM WITHDRAWAL AUTHORIZED ON MM/DD <location...>
    m = re.match(r"^ATM\s+WITHDRAWAL\s+AUTHORIZED\s+ON\s+\d{2}/\d{2}\s+(.*)$", d, flags=re.IGNORECASE)
    if m:
        return ("ATM WITHDRAWAL " + m.group(1).strip()).strip()

    # WIRE TRANS SVC CHARGE - SEQUENCE: <stuff>
    m = re.match(r"^WIRE\s+TRANS\s+SVC\s+CHARGE\s+-\s+SEQUENCE:\s+(.*)$", d, flags=re.IGNORECASE)
    if m:
        return ("WIRE TRANS SVC CHARGE " + m.group(1).strip()).strip()

    # WT FED#... <beneficiary...>
    m = re.match(r"^WT\s+FED#\S+\s+(.*)$", d, flags=re.IGNORECASE)
    if m:
        return ("WT FED " + m.group(1).strip()).strip()

    return d


def merchant_core(description: str) -> str:
    """
    Produce a merchant core string that matches similar merchants even when the
    bank adds store numbers, long spacing, etc.

    Examples:
      "7-ELEVEN 21494 AURORA, CO"           -> "7-ELEVEN"
      "COSTCO GAS #1652 DENVER, CO"         -> "COSTCO GAS"
      "AMAZON MKTPL*1O17I7S63 ..."          -> "AMAZON"
      "COMCAST-XFINITY  CABLE SVCS ..."     -> "COMCAST/XFINITY"
    """
    d = normalize_spaces(description).upper()
    if not d:
        return "OTHER"

    # Strong explicit families first
    if d.startswith("AMAZON"):
        return "AMAZON"
    if d.startswith("ZELLE TO"):
        # group_key handles "ZELLE - person" — keep as fallback safe
        return "ZELLE"
    if d.startswith("APPLE.COM/BILL"):
        return "APPLE.COM/BILL"
    if d.startswith("7-ELEVEN"):
        return "7-ELEVEN"
    if d.startswith("COSTCO GAS"):
        return "COSTCO GAS"
    if d.startswith("COSTCO WHSE") or d.startswith("COSTCO WHOLESALE"):
        return "COSTCO WHSE"
    if d.startswith("KING SOOPERS"):
        return "KING SOOPERS"
    if d.startswith("SPROUTS"):
        return "SPROUTS"
    if d.startswith("WAL-MART") or d.startswith("WM SUPERCENTER"):
        return "WALMART"
    if d.startswith("COMCAST") or d.startswith("COMCAST-XFINITY") or d.startswith("XFINITY"):
        return "COMCAST/XFINITY"
    if d.startswith("ATM WITHDRAWAL"):
        return "ATM WITHDRAWAL"
    if d.startswith("STATE FARM"):
        return "STATE FARM"
    if d.startswith("DEPT EDUCATION") or "STUDENT LN" in d:
        return "STUDENT LOAN"
    if d.startswith("PENNYMAC"):
        return "PENNYMAC"

    # Generic cleanup:
    tokens = d.split()
    if not tokens:
        return "OTHER"

    # If second token looks like a store number / code, keep first token
    if len(tokens) >= 2 and (tokens[1].isdigit() or re.fullmatch(r"#\d+", tokens[1])):
        return tokens[0]

    # Default: first two tokens (BEST BUY, CITY OF, ADVANCE AUTO, etc.)
    return " ".join(tokens[:2]) if len(tokens) >= 2 else tokens[0]


# -----------------------------
# Grouping logic
# -----------------------------
def extract_zelle_person(desc_upper: str) -> str:
    """
    Extract Zelle recipient from:
      "ZELLE TO ABAGEZ DANIEL ON 04/30 REF #..."
    Returns "UNKNOWN" if not found.
    """
    d = normalize_spaces(desc_upper)
    if not d.startswith("ZELLE TO"):
        return "UNKNOWN"

    rest = d[len("ZELLE TO"):].strip()

    if " ON " in rest:
        person = rest.split(" ON ", 1)[0].strip()
    elif " REF " in rest:
        person = rest.split(" REF ", 1)[0].strip()
    else:
        person = " ".join(rest.split()[:3]).strip()

    person = normalize_spaces(person)
    return person if person else "UNKNOWN"


def group_key(description: str) -> str:
    """
    Group / Family key rules:
    - ZELLE: split per person -> "ZELLE - <PERSON>"
    - Everything else: use merchant_core() so narration variations match.
    """
    d = normalize_spaces(description).upper()
    if not d:
        return "OTHER"

    if d.startswith("ZELLE TO"):
        return f"ZELLE - {extract_zelle_person(d)}"

    return merchant_core(d)


# -----------------------------
# CSV Load + Clean + Sort
# -----------------------------
def load_csv_rows(csv_path: Path):
    """Load CSV as list of dict rows plus headers."""
    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        headers = reader.fieldnames
    return headers, rows


def clean_rows(rows):
    """
    Clean rows:
    - normalize Description spacing
    - strip narration prefixes like PURCHASE AUTHORIZED ON ...
    - remove ONLINE TRANSFER REF...
    - normalize Payment Method
    Returns: cleaned_rows, removed_count
    """
    cleaned = []
    removed = 0

    for r in rows:
        r["Description"] = clean_description(r.get("Description"))

        if (r.get("Description") or "").upper().startswith(REMOVE_DESC_PREFIX.upper()):
            removed += 1
            continue

        r["Payment Method"] = normalize_payment_method(r.get("Payment Method"))
        cleaned.append(r)

    return cleaned, removed


def sort_rows_for_grouping(rows):
    """Sort by group -> description -> date (oldest first)."""
    rows.sort(
        key=lambda r: (
            group_key(r.get("Description") or ""),
            (r.get("Description") or "").upper(),
            parse_date(r.get("Date")) or datetime.max,
        )
    )
    return rows


# -----------------------------
# Excel Output: DETAIL (rows + subtotals)
# -----------------------------
def write_excel_detail_grouped(headers, rows, xlsx_path: Path):
    """
    Write Excel DETAIL:
    - all cleaned rows grouped by group_key()
    - adds TOTAL row + blank separator after each group
    """
    if not headers:
        raise ValueError("No headers detected in CSV.")
    if "Amount" not in headers or "Description" not in headers:
        raise ValueError("CSV must include 'Description' and 'Amount' columns.")

    amount_idx = headers.index("Amount") + 1
    desc_idx = headers.index("Description") + 1

    wb = Workbook()
    ws = wb.active
    ws.title = "Grouped Detail"

    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = BOLD

    def append_total(group_name, total_value, txn_count):
        row = [""] * len(headers)
        row[desc_idx - 1] = f"TOTAL ({group_name}) — {txn_count} txns"
        row[amount_idx - 1] = total_value
        ws.append(row)

        r = ws.max_row
        ws.cell(row=r, column=desc_idx).font = BOLD
        ws.cell(row=r, column=amount_idx).font = BOLD

        ws.append([""] * len(headers))  # blank separator

    current_group = None
    group_total = 0.0
    group_count = 0

    for r in rows:
        g = group_key(r.get("Description") or "")

        if current_group is not None and g != current_group:
            append_total(current_group, group_total, group_count)
            group_total = 0.0
            group_count = 0

        current_group = g
        group_total += parse_amount(r.get("Amount"))
        group_count += 1

        ws.append([r.get(h, "") for h in headers])

    if current_group is not None:
        append_total(current_group, group_total, group_count)

    # format Amount column
    for i in range(2, ws.max_row + 1):
        ws.cell(row=i, column=amount_idx).number_format = '"$"#,##0.00'

    wb.save(xlsx_path)


# -----------------------------
# Excel Output: SUMMARY (group -> txns + total)
# -----------------------------
def write_excel_summary_by_group(rows, xlsx_path: Path):
    """
    Write Excel SUMMARY:
      Group | Txns | Total
    Sorted:
      - Txns: high -> low
      - Group: A -> Z (ties)
      - Total: high -> low (final tie-break)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Family Summary"

    headers = ["Group", "Txns", "Total"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = BOLD

    summary = {}
    for r in rows:
        g = group_key(r.get("Description") or "")
        amt = parse_amount(r.get("Amount"))
        if g not in summary:
            summary[g] = {"txns": 0, "total": 0.0}
        summary[g]["txns"] += 1
        summary[g]["total"] += amt

    sorted_items = sorted(summary.items(), key=lambda kv: (-kv[1]["txns"], kv[0], -kv[1]["total"]))

    grand_txns = 0
    grand_total = 0.0

    for gname, info in sorted_items:
        ws.append([gname, info["txns"], info["total"]])
        grand_txns += info["txns"]
        grand_total += info["total"]

    ws.append(["GRAND TOTAL", grand_txns, grand_total])
    last_row = ws.max_row
    ws.cell(row=last_row, column=1).font = BOLD
    ws.cell(row=last_row, column=2).font = BOLD
    ws.cell(row=last_row, column=3).font = BOLD

    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=3).number_format = '"$"#,##0.00'

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 16

    wb.save(xlsx_path)


# -----------------------------
# PDF Output: DETAIL
# -----------------------------
def build_pdf_detail(pdf_path: Path, rows, removed_count: int):
    """Build a detailed PDF: each group with a table and total (one group per page)."""
    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(
        str(pdf_path),
        pagesize=letter,
        leftMargin=0.6 * inch,
        rightMargin=0.6 * inch,
        topMargin=0.6 * inch,
        bottomMargin=0.6 * inch,
    )

    story = []
    story.append(Paragraph("Expenses - Detailed Grouped Report", styles["Title"]))
    story.append(Spacer(1, 0.12 * inch))
    story.append(Paragraph(
        f"Removed rows starting with: <b>{REMOVE_DESC_PREFIX}</b> — <b>{removed_count}</b> removed.",
        styles["Normal"]
    ))
    story.append(Spacer(1, 0.18 * inch))

    groups = {}
    for r in rows:
        g = group_key(r.get("Description") or "")
        groups.setdefault(g, []).append(r)

    grand_total = 0.0

    for gname in sorted(groups.keys()):
        grows = groups[gname]
        gtotal = sum(parse_amount(r.get("Amount")) for r in grows)
        grand_total += gtotal

        story.append(Paragraph(
            f"<b>Group:</b> {gname} &nbsp;&nbsp; <b>Txns:</b> {len(grows)} &nbsp;&nbsp; <b>Total:</b> {fmt_money(gtotal)}",
            styles["Heading2"]
        ))
        story.append(Spacer(1, 0.08 * inch))

        table_data = [["Date", "Description", "Payee", "Payment Method", "Amount"]]
        for r in grows:
            table_data.append([
                (r.get("Date") or "").strip(),
                (r.get("Description") or "").strip(),
                (r.get("Payee") or "").strip(),
                (r.get("Payment Method") or "").strip(),
                fmt_money(parse_amount(r.get("Amount"))),
            ])

        tbl = Table(
            table_data,
            colWidths=[0.9 * inch, 3.1 * inch, 1.4 * inch, 1.6 * inch, 0.9 * inch],
            repeatRows=1
        )
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ALIGN", (-1, 1), (-1, -1), "RIGHT"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ]))

        story.append(tbl)
        story.append(PageBreak())

    story.append(Paragraph("Grand Total", styles["Heading1"]))
    story.append(Spacer(1, 0.1 * inch))
    story.append(Paragraph(f"<b>{fmt_money(grand_total)}</b>", styles["Normal"]))

    doc.build(story)


# -----------------------------
# PDF Output: SUMMARY
# -----------------------------
def build_pdf_summary(pdf_path: Path, rows, removed_count: int):
    """Build a summary PDF: Group | Txns | Total (sorted txns desc, A→Z ties)."""
    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(
        str(pdf_path),
        pagesize=letter,
        leftMargin=0.75 * inch,
        rightMargin=0.75 * inch,
        topMargin=0.75 * inch,
        bottomMargin=0.75 * inch,
    )

    summary = {}
    for r in rows:
        g = group_key(r.get("Description") or "")
        amt = parse_amount(r.get("Amount"))
        if g not in summary:
            summary[g] = {"txns": 0, "total": 0.0}
        summary[g]["txns"] += 1
        summary[g]["total"] += amt

    sorted_items = sorted(summary.items(), key=lambda kv: (-kv[1]["txns"], kv[0], -kv[1]["total"]))

    story = []
    story.append(Paragraph("Expense Summary by Group (Txns desc, A→Z ties)", styles["Title"]))
    story.append(Spacer(1, 0.15 * inch))
    story.append(Paragraph(
        f"Removed rows starting with: <b>{REMOVE_DESC_PREFIX}</b> — <b>{removed_count}</b> removed.",
        styles["Normal"]
    ))
    story.append(Spacer(1, 0.2 * inch))

    table_data = [["Group", "Txns", "Total"]]
    grand_total = 0.0
    total_txns = 0

    for gname, info in sorted_items:
        table_data.append([gname, str(info["txns"]), fmt_money(info["total"])])
        total_txns += info["txns"]
        grand_total += info["total"]

    table_data.append(["GRAND TOTAL", str(total_txns), fmt_money(grand_total)])

    tbl = Table(table_data, colWidths=[3.8 * inch, 0.8 * inch, 1.4 * inch], repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.whitesmoke),
    ]))

    story.append(tbl)
    doc.build(story)


# -----------------------------
# CLI / Main
# -----------------------------
def main():
    parser = argparse.ArgumentParser(description="Clean + Group + Export finance CSV to Excel/PDF.")
    parser.add_argument("--in", dest="input_csv", default=DEFAULT_INPUT_CSV, help="Input CSV filename (same folder).")

    parser.add_argument("--excel-detail", action="store_true", help="Generate Excel detail grouped output.")
    parser.add_argument("--excel-summary", action="store_true", help="Generate Excel summary (group, txns, total).")
    parser.add_argument("--pdf-detail", action="store_true", help="Generate detailed grouped PDF.")
    parser.add_argument("--pdf-summary", action="store_true", help="Generate summary PDF (txns/totals).")

    parser.add_argument("--excel-detail-out", default=DEFAULT_EXCEL_DETAIL_OUT, help="Excel detail output filename.")
    parser.add_argument("--excel-summary-out", default=DEFAULT_EXCEL_SUMMARY_OUT, help="Excel summary output filename.")
    parser.add_argument("--pdf-detail-out", default=DEFAULT_PDF_DETAIL_OUT, help="Detailed PDF output filename.")
    parser.add_argument("--pdf-summary-out", default=DEFAULT_PDF_SUMMARY_OUT, help="Summary PDF output filename.")
    args = parser.parse_args()

    # If no flags were provided, generate everything
    generate_all = not (args.excel_detail or args.excel_summary or args.pdf_detail or args.pdf_summary)

    do_excel_detail = args.excel_detail or generate_all
    do_excel_summary = args.excel_summary or generate_all
    do_pdf_detail = args.pdf_detail or generate_all
    do_pdf_summary = args.pdf_summary or generate_all

    base_dir = Path(__file__).parent
    csv_path = base_dir / args.input_csv

    if not csv_path.exists():
        raise FileNotFoundError(f"Input CSV not found: {csv_path}")

    headers, rows = load_csv_rows(csv_path)
    cleaned, removed = clean_rows(rows)
    sorted_rows = sort_rows_for_grouping(cleaned)

    if do_excel_detail:
        out_path = base_dir / args.excel_detail_out
        write_excel_detail_grouped(headers, sorted_rows, out_path)
        print(f"✅ Excel detail: {out_path.name}")

    if do_excel_summary:
        out_path = base_dir / args.excel_summary_out
        write_excel_summary_by_group(sorted_rows, out_path)
        print(f"✅ Excel summary: {out_path.name}")

    if do_pdf_detail:
        out_path = base_dir / args.pdf_detail_out
        build_pdf_detail(out_path, sorted_rows, removed)
        print(f"✅ PDF detail: {out_path.name}")

    if do_pdf_summary:
        out_path = base_dir / args.pdf_summary_out
        build_pdf_summary(out_path, sorted_rows, removed)
        print(f"✅ PDF summary: {out_path.name}")

    print("🎉 Done")


if __name__ == "__main__":
    main()
#!/usr/bin/env python3
"""
finance_pipeline.py (ONE DRY FILE)

Reads a bank-style expenses CSV and produces:
1) Excel DETAIL (.xlsx): cleaned + grouped rows with subtotal lines
2) Excel SUMMARY (.xlsx): Group | Txns | Total (sorted txns desc, A→Z ties)
3) PDF DETAIL (.pdf): grouped by group/family with tables
4) PDF SUMMARY (.pdf): Group | Txns | Total (sorted txns desc, A→Z ties)

Folder setup:
- Put this script and expenses.csv in the same folder.

Run:
  python3 finance_pipeline.py
  python3 finance_pipeline.py --excel-summary
  python3 finance_pipeline.py --excel-detail --pdf-summary
"""

import csv
import re
import argparse
from pathlib import Path
from datetime import datetime

# Excel
from openpyxl import Workbook
from openpyxl.styles import Font

# PDF
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
)


# -----------------------------
# Defaults / Config
# -----------------------------
DEFAULT_INPUT_CSV = "expenses_raw_spacing_fixed.csv" #

DEFAULT_EXCEL_DETAIL_OUT = "expenses_clean_grouped.xlsx"
DEFAULT_EXCEL_SUMMARY_OUT = "expenses_family_summary.xlsx"

DEFAULT_PDF_DETAIL_OUT = "expenses_grouped_families_detail.pdf"
DEFAULT_PDF_SUMMARY_OUT = "expenses_family_summary_by_txns_AZ.pdf"

REMOVE_DESC_PREFIX = "ONLINE TRANSFER REF"

WF_CARD_PREFIX = "WELLS FARGO ACTIVE CASH VISA(R) CARD"
WF_CARD_ALIAS = "WFACV"

DATE_FORMATS = (
    "%m/%d/%Y",
    "%m/%d/%y",
    "%Y-%m-%d",
    "%m-%d-%Y",
    "%m-%d-%y",
)

BOLD = Font(bold=True)


# -----------------------------
# Core Normalizers / Parsers
# -----------------------------
def normalize_spaces(text: str) -> str:
    """Trim and collapse internal whitespace to a single space."""
    return " ".join((text or "").split()).strip()


def normalize_payment_method(value: str) -> str:
    """
    Normalize WF Active Cash Visa payment method:
    'WELLS FARGO ACTIVE CASH VISA(R) CARD ...4321' -> 'WFACV...4321'
    """
    value = (value or "").strip()
    if value.upper().startswith(WF_CARD_PREFIX):
        suffix = value[len(WF_CARD_PREFIX):].strip()
        return f"{WF_CARD_ALIAS}{suffix}"
    return value


def parse_date(value: str):
    """Parse common bank/Excel date formats. Returns datetime or None."""
    s = ("" if value is None else str(value)).strip()
    if not s:
        return None

    # strip time parts if present
    s = s.split()[0]
    if "T" in s:
        s = s.split("T")[0]

    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def parse_amount(value) -> float:
    """Convert Amount cell to float safely. Handles $ , and (negative) style."""
    if value is None:
        return 0.0

    s = str(value).strip()
    if not s:
        return 0.0

    negative = False
    if s.startswith("(") and s.endswith(")"):
        negative = True
        s = s[1:-1].strip()

    s = s.replace("$", "").replace(",", "")
    try:
        n = float(s)
        return -n if negative else n
    except ValueError:
        return 0.0


def fmt_money(n: float) -> str:
    """Format float as currency string for PDF."""
    return f"${n:,.2f}"


# -----------------------------
# Grouping logic
# -----------------------------
def extract_zelle_person(desc_upper: str) -> str:
    """
    Extract Zelle recipient from:
      "ZELLE TO ABAGEZ DANIEL ON 04/30 REF #..."
    Returns "UNKNOWN" if not found.
    """
    d = normalize_spaces(desc_upper)
    if not d.startswith("ZELLE TO"):
        return "UNKNOWN"

    rest = d[len("ZELLE TO"):].strip()

    if " ON " in rest:
        person = rest.split(" ON ", 1)[0].strip()
    elif " REF " in rest:
        person = rest.split(" REF ", 1)[0].strip()
    else:
        person = " ".join(rest.split()[:3]).strip()

    person = normalize_spaces(person)
    return person if person else "UNKNOWN"


def group_key(description: str) -> str:
    """
    Group / Family key rules:
    - ZELLE: split per person -> "ZELLE - <PERSON>"
    - AMAZON: group all together -> "AMAZON"
    - Others: common families + fallback grouping
    """
    d = normalize_spaces(description).upper()
    if not d:
        return "OTHER"

    # Zelle per person
    if d.startswith("ZELLE TO"):
        return f"ZELLE - {extract_zelle_person(d)}"

    # Amazon all together
    if d.startswith("AMAZON"):
        return "AMAZON"

    # Common families in your dataset
    if d.startswith("ATM WITHDRAWAL"):
        return "ATM WITHDRAWAL"
    if d.startswith("COMCAST") or d.startswith("XFINITY"):
        return "COMCAST/XFINITY"
    if d.startswith("COSTCO GAS"):
        return "COSTCO GAS"
    if d.startswith("COSTCO WHSE") or d.startswith("COSTCO WHOLESALE"):
        return "COSTCO WHSE"
    if d.startswith("WAL-MART") or d.startswith("WM SUPERCENTER"):
        return "WALMART"
    if d.startswith("KING SOOPERS"):
        return "KING SOOPERS"
    if d.startswith("SPROUTS"):
        return "SPROUTS"
    if d.startswith("WHOLEFDS") or d.startswith("WHOLE FOODS"):
        return "WHOLE FOODS"
    if d.startswith("STATE FARM"):
        return "STATE FARM"
    if d.startswith("DEPT EDUCATION") or "STUDENT LN" in d:
        return "STUDENT LOAN"
    if d.startswith("PENNYMAC"):
        return "PENNYMAC"

    # Fallback grouping
    tokens = d.split()
    if not tokens:
        return "OTHER"

    # If second token is store id / code, group by first token
    if len(tokens) >= 2 and (tokens[1].isdigit() or re.fullmatch(r"#\d+", tokens[1] or "")):
        return tokens[0]

    # Else first two tokens helps group "BEST BUY", "CITY OF", etc.
    return " ".join(tokens[:2]) if len(tokens) >= 2 else tokens[0]


# -----------------------------
# CSV Load + Clean + Sort
# -----------------------------
def load_csv_rows(csv_path: Path):
    """Load CSV as list of dict rows plus headers."""
    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        headers = reader.fieldnames
    return headers, rows


def clean_rows(rows):
    """
    Clean rows:
    - normalize Description spacing
    - remove ONLINE TRANSFER REF...
    - normalize Payment Method
    Returns: cleaned_rows, removed_count
    """
    cleaned = []
    removed = 0

    for r in rows:
        r["Description"] = normalize_spaces(r.get("Description"))
        if (r.get("Description") or "").upper().startswith(REMOVE_DESC_PREFIX.upper()):
            removed += 1
            continue

        r["Payment Method"] = normalize_payment_method(r.get("Payment Method"))
        cleaned.append(r)

    return cleaned, removed


def sort_rows_for_grouping(rows):
    """Sort by group -> description -> date (oldest first)."""
    rows.sort(
        key=lambda r: (
            group_key(r.get("Description") or ""),
            (r.get("Description") or "").upper(),
            parse_date(r.get("Date")) or datetime.max,
        )
    )
    return rows


# -----------------------------
# Excel Output: DETAIL (rows + subtotals)
# -----------------------------
def write_excel_detail_grouped(headers, rows, xlsx_path: Path):
    """
    Write Excel DETAIL:
    - all cleaned rows grouped by group_key()
    - adds TOTAL row + blank separator after each group
    """
    if "Amount" not in headers or "Description" not in headers:
        raise ValueError("CSV must include 'Description' and 'Amount' columns.")

    amount_idx = headers.index("Amount") + 1
    desc_idx = headers.index("Description") + 1

    wb = Workbook()
    ws = wb.active
    ws.title = "Grouped Detail"

    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = BOLD

    def append_total(group_name, total_value, txn_count):
        row = [""] * len(headers)
        row[desc_idx - 1] = f"TOTAL ({group_name}) — {txn_count} txns"
        row[amount_idx - 1] = total_value
        ws.append(row)

        r = ws.max_row
        ws.cell(row=r, column=desc_idx).font = BOLD
        ws.cell(row=r, column=amount_idx).font = BOLD

        ws.append([""] * len(headers))  # blank separator

    current_group = None
    group_total = 0.0
    group_count = 0

    for r in rows:
        g = group_key(r.get("Description") or "")

        if current_group is not None and g != current_group:
            append_total(current_group, group_total, group_count)
            group_total = 0.0
            group_count = 0

        current_group = g
        group_total += parse_amount(r.get("Amount"))
        group_count += 1

        ws.append([r.get(h, "") for h in headers])

    if current_group is not None:
        append_total(current_group, group_total, group_count)

    # format Amount column
    for i in range(2, ws.max_row + 1):
        ws.cell(row=i, column=amount_idx).number_format = '"$"#,##0.00'

    wb.save(xlsx_path)


# -----------------------------
# Excel Output: SUMMARY (group -> txns + total)
# -----------------------------
def write_excel_summary_by_group(rows, xlsx_path: Path):
    """
    Write Excel SUMMARY:
      Group | Txns | Total
    Sorted:
      - Txns: high -> low
      - Group: A -> Z (ties)
      - Total: high -> low (final tie-break)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Family Summary"

    headers = ["Group", "Txns", "Total"]
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = BOLD

    # Build summary
    summary = {}
    for r in rows:
        g = group_key(r.get("Description") or "")
        amt = parse_amount(r.get("Amount"))
        if g not in summary:
            summary[g] = {"txns": 0, "total": 0.0}
        summary[g]["txns"] += 1
        summary[g]["total"] += amt

    # Sort: txns desc, group A→Z, total desc
    sorted_items = sorted(summary.items(), key=lambda kv: (-kv[1]["txns"], kv[0], -kv[1]["total"]))

    grand_txns = 0
    grand_total = 0.0

    for gname, info in sorted_items:
        ws.append([gname, info["txns"], info["total"]])
        grand_txns += info["txns"]
        grand_total += info["total"]

    # Grand total row
    ws.append(["GRAND TOTAL", grand_txns, grand_total])
    last_row = ws.max_row
    ws.cell(row=last_row, column=1).font = BOLD
    ws.cell(row=last_row, column=2).font = BOLD
    ws.cell(row=last_row, column=3).font = BOLD

    # Format Total column as currency
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=3).number_format = '"$"#,##0.00'

    # Friendly widths
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 16

    wb.save(xlsx_path)


# -----------------------------
# PDF Output: DETAIL
# -----------------------------
def build_pdf_detail(pdf_path: Path, rows, removed_count: int):
    """Build a detailed PDF: each group with a table and total (one group per page)."""
    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(
        str(pdf_path),
        pagesize=letter,
        leftMargin=0.6 * inch,
        rightMargin=0.6 * inch,
        topMargin=0.6 * inch,
        bottomMargin=0.6 * inch,
    )

    story = []
    story.append(Paragraph("Expenses - Detailed Grouped Report", styles["Title"]))
    story.append(Spacer(1, 0.12 * inch))
    story.append(Paragraph(
        f"Removed rows starting with: <b>{REMOVE_DESC_PREFIX}</b> — <b>{removed_count}</b> removed.",
        styles["Normal"]
    ))
    story.append(Spacer(1, 0.18 * inch))

    # group rows
    groups = {}
    for r in rows:
        g = group_key(r.get("Description") or "")
        groups.setdefault(g, []).append(r)

    grand_total = 0.0

    for gname in sorted(groups.keys()):
        grows = groups[gname]
        gtotal = sum(parse_amount(r.get("Amount")) for r in grows)
        grand_total += gtotal

        story.append(Paragraph(
            f"<b>Group:</b> {gname} &nbsp;&nbsp; <b>Txns:</b> {len(grows)} &nbsp;&nbsp; <b>Total:</b> {fmt_money(gtotal)}",
            styles["Heading2"]
        ))
        story.append(Spacer(1, 0.08 * inch))

        table_data = [["Date", "Description", "Payee", "Payment Method", "Amount"]]
        for r in grows:
            table_data.append([
                (r.get("Date") or "").strip(),
                (r.get("Description") or "").strip(),
                (r.get("Payee") or "").strip(),
                (r.get("Payment Method") or "").strip(),
                fmt_money(parse_amount(r.get("Amount"))),
            ])

        tbl = Table(
            table_data,
            colWidths=[0.9 * inch, 3.1 * inch, 1.4 * inch, 1.6 * inch, 0.9 * inch],
            repeatRows=1
        )
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ALIGN", (-1, 1), (-1, -1), "RIGHT"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ]))

        story.append(tbl)
        story.append(PageBreak())

    story.append(Paragraph("Grand Total", styles["Heading1"]))
    story.append(Spacer(1, 0.1 * inch))
    story.append(Paragraph(f"<b>{fmt_money(grand_total)}</b>", styles["Normal"]))

    doc.build(story)


# -----------------------------
# PDF Output: SUMMARY
# -----------------------------
def build_pdf_summary(pdf_path: Path, rows, removed_count: int):
    """
    Build a summary PDF:
    Group | Txns | Total
    Sorted: txns desc, group A→Z ties, total desc last tie-break
    """
    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(
        str(pdf_path),
        pagesize=letter,
        leftMargin=0.75 * inch,
        rightMargin=0.75 * inch,
        topMargin=0.75 * inch,
        bottomMargin=0.75 * inch,
    )

    summary = {}
    for r in rows:
        g = group_key(r.get("Description") or "")
        amt = parse_amount(r.get("Amount"))
        if g not in summary:
            summary[g] = {"txns": 0, "total": 0.0}
        summary[g]["txns"] += 1
        summary[g]["total"] += amt

    sorted_items = sorted(summary.items(), key=lambda kv: (-kv[1]["txns"], kv[0], -kv[1]["total"]))

    story = []
    story.append(Paragraph("Expense Summary by Group (Txns desc, A→Z ties)", styles["Title"]))
    story.append(Spacer(1, 0.15 * inch))
    story.append(Paragraph(
        f"Removed rows starting with: <b>{REMOVE_DESC_PREFIX}</b> — <b>{removed_count}</b> removed.",
        styles["Normal"]
    ))
    story.append(Spacer(1, 0.2 * inch))

    table_data = [["Group", "Txns", "Total"]]
    grand_total = 0.0
    total_txns = 0

    for gname, info in sorted_items:
        table_data.append([gname, str(info["txns"]), fmt_money(info["total"])])
        total_txns += info["txns"]
        grand_total += info["total"]

    table_data.append(["GRAND TOTAL", str(total_txns), fmt_money(grand_total)])

    tbl = Table(table_data, colWidths=[3.8 * inch, 0.8 * inch, 1.4 * inch], repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.whitesmoke),
    ]))

    story.append(tbl)
    doc.build(story)


# -----------------------------
# CLI / Main
# -----------------------------
def main():
    parser = argparse.ArgumentParser(description="Clean + Group + Export finance CSV to Excel/PDF.")
    parser.add_argument("--in", dest="input_csv", default=DEFAULT_INPUT_CSV, help="Input CSV filename (same folder).")

    parser.add_argument("--excel-detail", action="store_true", help="Generate Excel detail grouped output.")
    parser.add_argument("--excel-summary", action="store_true", help="Generate Excel summary (group, txns, total).")
    parser.add_argument("--pdf-detail", action="store_true", help="Generate detailed grouped PDF.")
    parser.add_argument("--pdf-summary", action="store_true", help="Generate summary PDF (txns/totals).")

    parser.add_argument("--excel-detail-out", default=DEFAULT_EXCEL_DETAIL_OUT, help="Excel detail output filename.")
    parser.add_argument("--excel-summary-out", default=DEFAULT_EXCEL_SUMMARY_OUT, help="Excel summary output filename.")
    parser.add_argument("--pdf-detail-out", default=DEFAULT_PDF_DETAIL_OUT, help="Detailed PDF output filename.")
    parser.add_argument("--pdf-summary-out", default=DEFAULT_PDF_SUMMARY_OUT, help="Summary PDF output filename.")
    args = parser.parse_args()

    # If no flags were provided, generate everything
    generate_all = not (args.excel_detail or args.excel_summary or args.pdf_detail or args.pdf_summary)

    do_excel_detail = args.excel_detail or generate_all
    do_excel_summary = args.excel_summary or generate_all
    do_pdf_detail = args.pdf_detail or generate_all
    do_pdf_summary = args.pdf_summary or generate_all

    base_dir = Path(__file__).parent
    csv_path = base_dir / args.input_csv

    if not csv_path.exists():
        raise FileNotFoundError(f"Input CSV not found: {csv_path}")

    headers, rows = load_csv_rows(csv_path)
    cleaned, removed = clean_rows(rows)
    sorted_rows = sort_rows_for_grouping(cleaned)

    if do_excel_detail:
        out_path = base_dir / args.excel_detail_out
        write_excel_detail_grouped(headers, sorted_rows, out_path)
        print(f"✅ Excel detail: {out_path.name}")

    if do_excel_summary:
        out_path = base_dir / args.excel_summary_out
        write_excel_summary_by_group(sorted_rows, out_path)
        print(f"✅ Excel summary: {out_path.name}")

    if do_pdf_detail:
        out_path = base_dir / args.pdf_detail_out
        build_pdf_detail(out_path, sorted_rows, removed)
        print(f"✅ PDF detail: {out_path.name}")

    if do_pdf_summary:
        out_path = base_dir / args.pdf_summary_out
        build_pdf_summary(out_path, sorted_rows, removed)
        print(f"✅ PDF summary: {out_path.name}")

    print("🎉 Done")


if __name__ == "__main__":
    main()
#!/usr/bin/env python3
"""
finance_pipeline.py  (ONE DRY FILE)

Reads a Wells Fargo-style expenses CSV and produces:
1) Excel (.xlsx) cleaned + sorted + grouped with subtotals
2) Detailed PDF grouped by families (each family on its own page)
3) PDF summary (group/person) -> txns + totals (sorted: txns desc, name A→Z ties)

Put this file in the SAME folder as:
- expenses.csv

Run examples:
  python3 finance_pipeline.py
  python3 finance_pipeline.py --in expenses.csv --excel --pdf-detail --pdf-summary
"""

import csv
import re
import argparse
from pathlib import Path
from datetime import datetime

# Excel
from openpyxl import Workbook
from openpyxl.styles import Font

# PDF
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
)


# -----------------------------
# Defaults / Config
# -----------------------------
DEFAULT_INPUT_CSV = "expenses_raw_spacing_fixed.csv" #

DEFAULT_EXCEL_OUT = "expenses_clean_grouped.xlsx"
DEFAULT_PDF_DETAIL_OUT = "expenses_grouped_families_detail.pdf"
DEFAULT_PDF_SUMMARY_OUT = "expenses_family_summary_by_txns_AZ.pdf"

REMOVE_DESC_PREFIX = "ONLINE TRANSFER REF"

WF_CARD_PREFIX = "WELLS FARGO ACTIVE CASH VISA(R) CARD"
WF_CARD_ALIAS = "WFACV"

DATE_FORMATS = (
    "%m/%d/%Y",
    "%m/%d/%y",
    "%Y-%m-%d",
    "%m-%d-%Y",
    "%m-%d-%y",
)

BOLD = Font(bold=True)


# -----------------------------
# Core Normalizers / Parsers
# -----------------------------
def normalize_spaces(text: str) -> str:
    """Trim and collapse internal whitespace to a single space."""
    return " ".join((text or "").split()).strip()


def normalize_payment_method(value: str) -> str:
    """
    Normalize WF Active Cash Visa payment method:
    'WELLS FARGO ACTIVE CASH VISA(R) CARD ...4321' -> 'WFACV...4321'
    """
    value = (value or "").strip()
    if value.upper().startswith(WF_CARD_PREFIX):
        suffix = value[len(WF_CARD_PREFIX):].strip()
        return f"{WF_CARD_ALIAS}{suffix}"
    return value


def parse_date(value: str):
    """
    Parse common bank/Excel date formats. Returns datetime or None.
    Unparseable dates can be handled by callers (e.g., sort to end).
    """
    s = ("" if value is None else str(value)).strip()
    if not s:
        return None

    # strip time parts if present
    s = s.split()[0]
    if "T" in s:
        s = s.split("T")[0]

    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def parse_amount(value) -> float:
    """
    Convert Amount cell to float safely.
    Handles: "$1,234.56", "1234.56", "(123.45)"
    """
    if value is None:
        return 0.0

    s = str(value).strip()
    if not s:
        return 0.0

    negative = False
    if s.startswith("(") and s.endswith(")"):
        negative = True
        s = s[1:-1].strip()

    s = s.replace("$", "").replace(",", "")
    try:
        n = float(s)
        return -n if negative else n
    except ValueError:
        return 0.0


def fmt_money(n: float) -> str:
    """Format float as currency string for PDF."""
    return f"${n:,.2f}"


# -----------------------------
# Grouping logic
# -----------------------------
def extract_zelle_person(desc_upper: str) -> str:
    """
    Extract Zelle recipient from:
      "ZELLE TO ABAGEZ DANIEL ON 04/30 REF #..."
    Returns "UNKNOWN" if not found.
    """
    d = normalize_spaces(desc_upper)
    if not d.startswith("ZELLE TO"):
        return "UNKNOWN"

    rest = d[len("ZELLE TO"):].strip()

    if " ON " in rest:
        person = rest.split(" ON ", 1)[0].strip()
    elif " REF " in rest:
        person = rest.split(" REF ", 1)[0].strip()
    else:
        person = " ".join(rest.split()[:3]).strip()

    person = normalize_spaces(person)
    return person if person else "UNKNOWN"


def group_key(description: str) -> str:
    """
    Group / Family key rules:
    - ZELLE: split per person -> "ZELLE - <PERSON>"
    - AMAZON: group all together -> "AMAZON"
    - Others: common families + fallback grouping
    """
    d = normalize_spaces(description).upper()
    if not d:
        return "OTHER"

    # Zelle per person
    if d.startswith("ZELLE TO"):
        return f"ZELLE - {extract_zelle_person(d)}"

    # Amazon all together
    if d.startswith("AMAZON"):
        return "AMAZON"

    # Helpful families from your dataset
    if d.startswith("ATM WITHDRAWAL"):
        return "ATM WITHDRAWAL"
    if d.startswith("COMCAST") or d.startswith("XFINITY"):
        return "COMCAST/XFINITY"
    if d.startswith("COSTCO GAS"):
        return "COSTCO GAS"
    if d.startswith("COSTCO WHSE") or d.startswith("COSTCO WHOLESALE"):
        return "COSTCO WHSE"
    if d.startswith("WAL-MART") or d.startswith("WM SUPERCENTER"):
        return "WALMART"
    if d.startswith("KING SOOPERS"):
        return "KING SOOPERS"
    if d.startswith("SPROUTS"):
        return "SPROUTS"
    if d.startswith("WHOLEFDS") or d.startswith("WHOLE FOODS"):
        return "WHOLE FOODS"
    if d.startswith("STATE FARM"):
        return "STATE FARM"
    if d.startswith("DEPT EDUCATION") or "STUDENT LN" in d:
        return "STUDENT LOAN"
    if d.startswith("PENNYMAC"):
        return "PENNYMAC"

    # Fallback grouping
    tokens = d.split()
    if not tokens:
        return "OTHER"

    # If second token is store id / code, group by first token
    if len(tokens) >= 2 and (tokens[1].isdigit() or re.fullmatch(r"#\d+", tokens[1] or "")):
        return tokens[0]

    # Else first two tokens helps group "BEST BUY", "CITY OF", etc.
    return " ".join(tokens[:2]) if len(tokens) >= 2 else tokens[0]


# -----------------------------
# CSV Load + Clean
# -----------------------------
def load_csv_rows(csv_path: Path):
    """Load CSV as list of dict rows plus headers."""
    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        headers = reader.fieldnames
    return headers, rows


def clean_rows(rows):
    """
    Clean rows:
    - normalize Description spacing
    - remove ONLINE TRANSFER REF...
    - normalize Payment Method
    Returns: cleaned_rows, removed_count
    """
    cleaned = []
    removed = 0

    for r in rows:
        r["Description"] = normalize_spaces(r.get("Description"))
        if (r.get("Description") or "").upper().startswith(REMOVE_DESC_PREFIX.upper()):
            removed += 1
            continue

        r["Payment Method"] = normalize_payment_method(r.get("Payment Method"))
        cleaned.append(r)

    return cleaned, removed


def sort_rows_for_grouping(rows):
    """Sort by group -> description -> date (oldest first)."""
    rows.sort(
        key=lambda r: (
            group_key(r.get("Description") or ""),
            (r.get("Description") or "").upper(),
            parse_date(r.get("Date")) or datetime.max,
        )
    )
    return rows


# -----------------------------
# Excel Output (grouped + totals)
# -----------------------------
def write_excel_grouped(headers, rows, xlsx_path: Path):
    """
    Write Excel:
    - rows grouped by group_key()
    - adds TOTAL row + blank separator after each group
    """
    if "Amount" not in headers or "Description" not in headers:
        raise ValueError("CSV must include 'Description' and 'Amount' columns.")

    amount_idx = headers.index("Amount") + 1
    desc_idx = headers.index("Description") + 1

    wb = Workbook()
    ws = wb.active
    ws.title = "Grouped Totals"

    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = BOLD

    def append_total(group_name, total_value, txn_count):
        row = [""] * len(headers)
        row[desc_idx - 1] = f"TOTAL ({group_name}) — {txn_count} txns"
        row[amount_idx - 1] = total_value
        ws.append(row)

        r = ws.max_row
        ws.cell(row=r, column=desc_idx).font = BOLD
        ws.cell(row=r, column=amount_idx).font = BOLD

        ws.append([""] * len(headers))  # blank separator

    current_group = None
    group_total = 0.0
    group_count = 0

    for r in rows:
        g = group_key(r.get("Description") or "")

        if current_group is not None and g != current_group:
            append_total(current_group, group_total, group_count)
            group_total = 0.0
            group_count = 0

        current_group = g
        group_total += parse_amount(r.get("Amount"))
        group_count += 1

        ws.append([r.get(h, "") for h in headers])

    if current_group is not None:
        append_total(current_group, group_total, group_count)

    # format Amount column
    for i in range(2, ws.max_row + 1):
        ws.cell(row=i, column=amount_idx).number_format = "#,##0.00"

    wb.save(xlsx_path)


# -----------------------------
# PDF Output: Detailed grouped
# -----------------------------
def build_pdf_detail(pdf_path: Path, rows, removed_count: int):
    """Build a detailed PDF: each group with a table and subtotal."""
    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(
        str(pdf_path),
        pagesize=letter,
        leftMargin=0.6 * inch,
        rightMargin=0.6 * inch,
        topMargin=0.6 * inch,
        bottomMargin=0.6 * inch,
    )

    story = []
    story.append(Paragraph("Expenses - Detailed Grouped Report", styles["Title"]))
    story.append(Spacer(1, 0.12 * inch))
    story.append(Paragraph(
        f"Removed rows starting with: <b>{REMOVE_DESC_PREFIX}</b> — <b>{removed_count}</b> removed.",
        styles["Normal"]
    ))
    story.append(Spacer(1, 0.18 * inch))

    # group rows
    groups = {}
    for r in rows:
        g = group_key(r.get("Description") or "")
        groups.setdefault(g, []).append(r)

    grand_total = 0.0

    for gname in sorted(groups.keys()):
        grows = groups[gname]
        gtotal = sum(parse_amount(r.get("Amount")) for r in grows)
        grand_total += gtotal

        story.append(Paragraph(
            f"<b>Group:</b> {gname} &nbsp;&nbsp; <b>Txns:</b> {len(grows)} &nbsp;&nbsp; <b>Total:</b> {fmt_money(gtotal)}",
            styles["Heading2"]
        ))
        story.append(Spacer(1, 0.08 * inch))

        # Table columns (safe + readable)
        table_data = [["Date", "Description", "Payee", "Payment Method", "Amount"]]
        for r in grows:
            table_data.append([
                (r.get("Date") or "").strip(),
                (r.get("Description") or "").strip(),
                (r.get("Payee") or "").strip(),
                (r.get("Payment Method") or "").strip(),
                fmt_money(parse_amount(r.get("Amount"))),
            ])

        tbl = Table(
            table_data,
            colWidths=[0.9 * inch, 3.1 * inch, 1.4 * inch, 1.6 * inch, 0.9 * inch],
            repeatRows=1
        )
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ALIGN", (-1, 1), (-1, -1), "RIGHT"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
        ]))
        story.append(tbl)
        story.append(PageBreak())

    story.append(Paragraph("Grand Total", styles["Heading1"]))
    story.append(Spacer(1, 0.1 * inch))
    story.append(Paragraph(f"<b>{fmt_money(grand_total)}</b>", styles["Normal"]))

    doc.build(story)


# -----------------------------
# PDF Output: Summary
# -----------------------------
def build_pdf_summary(pdf_path: Path, rows, removed_count: int):
    """
    Build a summary PDF:
    Group name | Txns | Total
    Sorted: txns desc, group name A→Z ties, total desc last tie-break
    """
    styles = getSampleStyleSheet()
    doc = SimpleDocTemplate(
        str(pdf_path),
        pagesize=letter,
        leftMargin=0.75 * inch,
        rightMargin=0.75 * inch,
        topMargin=0.75 * inch,
        bottomMargin=0.75 * inch,
    )

    # summarize
    summary = {}
    for r in rows:
        g = group_key(r.get("Description") or "")
        amt = parse_amount(r.get("Amount"))
        if g not in summary:
            summary[g] = {"txns": 0, "total": 0.0}
        summary[g]["txns"] += 1
        summary[g]["total"] += amt

    # sort: txns desc, group name A→Z, total desc
    sorted_items = sorted(
        summary.items(),
        key=lambda kv: (-kv[1]["txns"], kv[0], -kv[1]["total"])
    )

    story = []
    story.append(Paragraph("Expense Summary by Group (Txns desc, A→Z ties)", styles["Title"]))
    story.append(Spacer(1, 0.15 * inch))
    story.append(Paragraph(
        f"Removed rows starting with: <b>{REMOVE_DESC_PREFIX}</b> — <b>{removed_count}</b> removed.",
        styles["Normal"]
    ))
    story.append(Spacer(1, 0.2 * inch))

    table_data = [["Group", "Txns", "Total"]]
    grand_total = 0.0
    total_txns = 0

    for gname, info in sorted_items:
        table_data.append([gname, str(info["txns"]), fmt_money(info["total"])])
        total_txns += info["txns"]
        grand_total += info["total"]

    table_data.append(["GRAND TOTAL", str(total_txns), fmt_money(grand_total)])

    tbl = Table(table_data, colWidths=[3.8 * inch, 0.8 * inch, 1.4 * inch], repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.whitesmoke),
    ]))

    story.append(tbl)
    doc.build(story)


# -----------------------------
# CLI / Main
# -----------------------------
def main():
    parser = argparse.ArgumentParser(description="Clean + Group + Export finance CSV to Excel/PDF.")
    parser.add_argument("--in", dest="input_csv", default=DEFAULT_INPUT_CSV, help="Input CSV filename (same folder).")
    parser.add_argument("--excel", action="store_true", help="Generate Excel grouped output.")
    parser.add_argument("--pdf-detail", action="store_true", help="Generate detailed grouped PDF.")
    parser.add_argument("--pdf-summary", action="store_true", help="Generate summary PDF (txns/totals).")
    parser.add_argument("--excel-out", default=DEFAULT_EXCEL_OUT, help="Excel output filename.")
    parser.add_argument("--pdf-detail-out", default=DEFAULT_PDF_DETAIL_OUT, help="Detailed PDF output filename.")
    parser.add_argument("--pdf-summary-out", default=DEFAULT_PDF_SUMMARY_OUT, help="Summary PDF output filename.")
    args = parser.parse_args()

    # If user gave no flags, generate all outputs
    generate_all = not (args.excel or args.pdf_detail or args.pdf_summary)
    do_excel = args.excel or generate_all
    do_pdf_detail = args.pdf_detail or generate_all
    do_pdf_summary = args.pdf_summary or generate_all

    base_dir = Path(__file__).parent
    csv_path = base_dir / args.input_csv

    if not csv_path.exists():
        raise FileNotFoundError(f"Input CSV not found: {csv_path}")

    headers, rows = load_csv_rows(csv_path)
    cleaned, removed = clean_rows(rows)
    sorted_rows = sort_rows_for_grouping(cleaned)

    if do_excel:
        xlsx_path = base_dir / args.excel_out
        write_excel_grouped(headers, sorted_rows, xlsx_path)
        print(f"✅ Excel: {xlsx_path.name}")

    if do_pdf_detail:
        pdf_path = base_dir / args.pdf_detail_out
        build_pdf_detail(pdf_path, sorted_rows, removed)
        print(f"✅ PDF detail: {pdf_path.name}")

    if do_pdf_summary:
        pdf_path = base_dir / args.pdf_summary_out
        build_pdf_summary(pdf_path, sorted_rows, removed)
        print(f"✅ PDF summary: {pdf_path.name}")

    print("🎉 Done")


if __name__ == "__main__":
    main()
