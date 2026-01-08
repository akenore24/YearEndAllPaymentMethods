#!/usr/bin/env python3
"""
finance_master.py — Single-file / DRY CLI (CSV -> Excel/PDF)

✅ Includes:
- output/ folder auto-created (csv/xlsx/pdf)
- merchant normalization (your merges: SHEGER MARKET, APPLEBEES, CHIPOTLE, DOMINO'S PIZZA, KING SOOPERS, NAME-CHEAP.COM, PRIMELENDING, WT FED, etc.)
- MT timestamp at top of Excel/PDF outputs
- ready_to_print (xlsx + pdf)
- quick_pdf_18mo (one executive summary PDF with 4 time buckets)
- NEW: exec_txns_desc (1-page executive summary sorted by txns high -> low)
- all (runs pipeline + ready_to_print + quick_pdf + quick_pdf_18mo + exec_txns_desc)

Install:
  pip3 install openpyxl reportlab

Run examples:
  python3 finance_master.py spacing
  python3 finance_master.py quick --limit 40
  python3 finance_master.py quick_pdf --limit 60 --sort total
  python3 finance_master.py exec_txns_desc --limit 25
  python3 finance_master.py quick_pdf_18mo --limit 15 --sort total --organized
  python3 finance_master.py pipeline
  python3 finance_master.py pdf_families --sort total --zelle-block first
  python3 finance_master.py excel_families --sort txns --zelle-block last
  python3 finance_master.py organized_pdf --top-total 25
  python3 finance_master.py ready_to_print --top-other 25
  python3 finance_master.py all

Use a different CSV:
  python3 finance_master.py --in transactions.csv all

Input CSV minimum columns:
  Date, Description, Amount
Optional:
  Payee, Payment Method, Location, Master Category, Subcategory
"""

from __future__ import annotations

import argparse
import csv
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional, Tuple

# -----------------------------
# Config
# -----------------------------
DEFAULT_INPUT_CSV = "expenses.csv"

# outputs (filenames)
DEFAULT_SPACING_OUT = "expenses_raw_spacing_fixed.csv"

DEFAULT_EXCEL_DETAIL_OUT = "expenses_clean_grouped.xlsx"
DEFAULT_EXCEL_SUMMARY_OUT = "expenses_family_summary.xlsx"
DEFAULT_EXCEL_FAMILIES_OUT = "expenses_family_summary_sorted.xlsx"

DEFAULT_PDF_DETAIL_OUT = "expenses_grouped_families_detail.pdf"
DEFAULT_PDF_SUMMARY_OUT = "expenses_family_summary.pdf"
DEFAULT_PDF_FAMILIES_SORTED_OUT = "expenses_family_totals_sorted.pdf"

DEFAULT_PDF_QUICK_OUT = "expenses_quick_summary.pdf"
DEFAULT_PDF_ORGANIZED_OUT = "organized_report.pdf"

READY_TO_PRINT_XLSX = "ready_to_print.xlsx"
READY_TO_PRINT_PDF = "ready_to_print.pdf"

# NEW: 18-month executive summary output
DEFAULT_PDF_QUICK_18MO_OUT = "expenses_quick_summary_18mo.pdf"

# NEW: exec txns high->low output
DEFAULT_PDF_HIGHEST_TXNS_OUT = "exec_txns_desc.pdf"

# folders
OUTPUT_DIR = Path("output")
OUT_CSV_DIR = OUTPUT_DIR / "csv"
OUT_XLSX_DIR = OUTPUT_DIR / "xlsx"
OUT_PDF_DIR = OUTPUT_DIR / "pdf"

# cleaning behavior
REMOVE_DESC_PREFIX = "ONLINE TRANSFER REF"

WF_CARD_PREFIX = "WELLS FARGO ACTIVE CASH VISA(R) CARD"
WF_CARD_ALIAS = "WFACV"

DATE_FORMATS: Tuple[str, ...] = (
    "%m/%d/%Y",
    "%m/%d/%y",
    "%Y-%m-%d",
    "%m-%d-%Y",
    "%m-%d-%y",
)

# Your requested 18-month buckets (explicit windows)
BUCKETS_18MO: List[Tuple[str, datetime, datetime]] = [
    ("1.0–3 months: Oct 1, 2025 – Dec 31, 2025", datetime(2025, 10, 1), datetime(2025, 12, 31)),
    ("4–6 months: Jul 1, 2025 – Sep 30, 2025", datetime(2025, 7, 1), datetime(2025, 9, 30)),
    ("7–12 months: Jan 1, 2025 – Jun 30, 2025", datetime(2025, 1, 1), datetime(2025, 6, 30)),
    ("13–18 months: Jul 1, 2024 – Dec 31, 2024", datetime(2024, 7, 1), datetime(2024, 12, 31)),
]

# priority pinned families for ready_to_print (ZELLE is unified in family view)
READY_FAMILIES_PRIORITY: List[str] = [
    "COSTCO WHSE",
    "COSTCO GAS",
    "ONLINE TRANSFER",
    "SHEGER MARKET",
    "PIASSA ETHIO",
    "WALMART",
    "ZELLE",
]

# Merchant normalization rules happen BEFORE grouping
# NOTE: normalize_merchant_name() uppercases before applying these rules.
MERCHANT_NORMALIZATION_RULES: List[Tuple[str, str]] = [
    # Noise suffix / ids
    (r"\bWT\s+FED[#\s]*\d+\b", "WT FED"),
    (r"\bEUNIFYPAY\*\s*", "EUNIFYPAY "),

    # Your requested merges
    (r"\bSHEGER\s+INTERNATION(?:AL)?\b", "SHEGER MARKET"),
    (r"\bAPPLEBEES\s+\d+\b", "APPLEBEES"),
    (r"\bCHIPOTLE\s+\d+\b", "CHIPOTLE"),
    (r"\bDOMINO['’]S\s+\d+\b", "DOMINO'S PIZZA"),
    (r"\bKING\s+SOOP(?:ERS)?\b", "KING SOOPERS"),
    (r"\bNAME-?CHEAP\.COM\s+[A-Z0-9]+\b", "NAME-CHEAP.COM"),
    (r"\bPRMG\s+WEB\b", "PRIMELENDING"),
    (r"\bPRIMELENDING\s+ACH\b", "PRIMELENDING"),
    (r"\bPRIMELENDING\s+WWW\.PRIMELEND,?TX\b", "PRIMELENDING"),
]

# -----------------------------
# Dependency helpers
# -----------------------------
def require_openpyxl():
    try:
        from openpyxl import Workbook  # noqa
        from openpyxl.styles import Font  # noqa
        return Workbook, Font
    except Exception:
        raise SystemExit("Missing dependency: openpyxl\nInstall with: pip3 install openpyxl\n")


def require_reportlab():
    try:
        from reportlab.lib.pagesizes import letter  # noqa
        from reportlab.lib.units import inch  # noqa
        from reportlab.lib import colors  # noqa
        from reportlab.lib.styles import getSampleStyleSheet  # noqa
        from reportlab.platypus import (  # noqa
            SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
        )
        return (letter, inch, colors, getSampleStyleSheet,
                SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak)
    except Exception:
        raise SystemExit("Missing dependency: reportlab\nInstall with: pip3 install reportlab\n")


# -----------------------------
# Output paths
# -----------------------------
def ensure_output_dirs() -> None:
    OUT_CSV_DIR.mkdir(parents=True, exist_ok=True)
    OUT_XLSX_DIR.mkdir(parents=True, exist_ok=True)
    OUT_PDF_DIR.mkdir(parents=True, exist_ok=True)


def out_path(kind: str, filename: str) -> Path:
    ensure_output_dirs()
    k = kind.lower()
    if k == "csv":
        return OUT_CSV_DIR / filename
    if k == "xlsx":
        return OUT_XLSX_DIR / filename
    if k == "pdf":
        return OUT_PDF_DIR / filename
    raise ValueError(f"Unknown output kind: {kind}")


# -----------------------------
# General utilities
# -----------------------------
def normalize_spaces(text: str) -> str:
    return " ".join((text or "").split()).strip()


def fmt_money(n: float) -> str:
    return f"${n:,.2f}"


def now_mountain() -> datetime:
    try:
        from zoneinfo import ZoneInfo
        return datetime.now(ZoneInfo("America/Denver"))
    except Exception:
        return datetime.now()


def mt_timestamp_line(prefix: str = "Generated") -> str:
    dt = now_mountain()
    return f"{prefix}: {dt.strftime('%Y-%m-%d %H:%M:%S')} MT"


# -----------------------------
# Parsing
# -----------------------------
def parse_amount(value) -> float:
    if value is None:
        return 0.0
    s = str(value).strip()
    if not s:
        return 0.0

    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg = True
        s = s[1:-1].strip()

    s = s.replace("$", "").replace(",", "")
    try:
        n = float(s)
        return -n if neg else n
    except ValueError:
        return 0.0


def parse_date(value: str) -> Optional[datetime]:
    s = ("" if value is None else str(value)).strip()
    if not s:
        return None
    s = s.split()[0]
    if "T" in s:
        s = s.split("T")[0]
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def normalize_payment_method(value: str) -> str:
    value = (value or "").strip()
    if value.upper().startswith(WF_CARD_PREFIX):
        suffix = value[len(WF_CARD_PREFIX):].strip()
        return f"{WF_CARD_ALIAS}{suffix}"
    return value


def normalize_merchant_name(description: str) -> str:
    """
    Merchant normalization layer:
    - uppercases
    - applies regex merge rules
    - strips common noise (asterisks, trailing ids, etc.)
    """
    if not description:
        return ""
    d = normalize_spaces(description).upper()

    # Apply explicit merge rules first
    for pattern, repl in MERCHANT_NORMALIZATION_RULES:
        d = re.sub(pattern, repl, d)

    # LYFT variants -> LYFT
    if d.startswith("LYFT"):
        d = "LYFT"

    # Generic cleanup
    d = re.sub(r"\*+", " ", d)
    d = re.sub(r"#\d+\b", "", d)         # remove trailing #digits tokens
    d = re.sub(r"\s+\d+\b$", "", d)      # remove trailing numeric store ids
    d = normalize_spaces(d)
    return d


# -----------------------------
# CSV IO
# -----------------------------
def load_csv_rows(csv_path: Path) -> Tuple[List[str], List[Dict[str, Any]]]:
    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        headers = reader.fieldnames or []
    return headers, rows


def write_csv_rows(out_path_: Path, headers: List[str], rows: List[Dict[str, Any]]) -> None:
    with open(out_path_, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=headers)
        w.writeheader()
        w.writerows(rows)


def ensure_required(headers: List[str], required: List[str]) -> None:
    missing = [h for h in required if h not in headers]
    if missing:
        raise ValueError(f"CSV missing required columns: {missing}")


# -----------------------------
# Cleaning
# -----------------------------
def clean_description(raw: str) -> str:
    d = normalize_spaces(raw)
    if not d:
        return ""

    m = re.match(r"^PURCHASE\s+AUTHORIZED\s+ON\s+\d{2}/\d{2}\s+(.*)$", d, flags=re.IGNORECASE)
    if m:
        return m.group(1).strip()

    m = re.match(r"^ATM\s+WITHDRAWAL\s+AUTHORIZED\s+ON\s+\d{2}/\d{2}\s+(.*)$", d, flags=re.IGNORECASE)
    if m:
        return ("ATM WITHDRAWAL " + m.group(1).strip()).strip()

    if re.match(r"^DEPOSITED\s+OR\s+CASHED\s+CHECK", d, flags=re.IGNORECASE):
        return "DEPOSITED OR CASHED CHECK"

    return d


def clean_rows(rows: List[Dict[str, Any]]) -> Tuple[List[Dict[str, Any]], int]:
    """
    - normalize Description narration + spacing
    - remove rows starting with ONLINE TRANSFER REF
    - normalize payment method
    """
    cleaned: List[Dict[str, Any]] = []
    removed = 0
    for r in rows:
        r["Description"] = clean_description(r.get("Description"))
        if (r.get("Description") or "").upper().startswith(REMOVE_DESC_PREFIX.upper()):
            removed += 1
            continue
        r["Payment Method"] = normalize_payment_method(r.get("Payment Method"))
        cleaned.append(r)
    return cleaned, removed


# -----------------------------
# Grouping
# -----------------------------
def extract_zelle_person(desc_upper: str) -> str:
    d = normalize_spaces(desc_upper)
    if not d.startswith("ZELLE TO"):
        return "UNKNOWN"
    rest = d[len("ZELLE TO"):].strip()
    if " ON " in rest:
        person = rest.split(" ON ", 1)[0].strip()
    elif " REF " in rest:
        person = rest.split(" REF ", 1)[0].strip()
    else:
        person = " ".join(rest.split()[:3]).strip()
    return normalize_spaces(person) or "UNKNOWN"


def merchant_core(description_upper: str) -> str:
    """
    Stable merchant family core (non-Zelle).
    Note: description_upper is already normalized via normalize_merchant_name().
    """
    d = description_upper
    if not d:
        return "OTHER"

    if d.startswith("AMAZON"):
        return "AMAZON"
    if d.startswith("7-ELEVEN"):
        return "7-ELEVEN"
    if d.startswith("COSTCO GAS"):
        return "COSTCO GAS"
    if d.startswith("COSTCO WHSE") or d.startswith("COSTCO WHOLESALE"):
        return "COSTCO WHSE"
    if d.startswith("WAL-MART") or d.startswith("WM SUPERCENTER"):
        return "WALMART"
    if d.startswith("KING SOOPERS"):
        return "KING SOOPERS"
    if d.startswith("SPROUTS"):
        return "SPROUTS"
    if d.startswith("WHOLEFDS") or d.startswith("WHOLE FOODS"):
        return "WHOLE FOODS"
    if d.startswith("COMCAST") or d.startswith("XFINITY"):
        return "COMCAST/XFINITY"
    if d.startswith("APPLE.COM/BILL"):
        return "APPLE.COM/BILL"
    if d.startswith("STATE FARM"):
        return "STATE FARM"
    if d.startswith("ATM WITHDRAWAL"):
        return "ATM WITHDRAWAL"
    if d.startswith("DEPT EDUCATION") or "STUDENT LN" in d:
        return "STUDENT LOAN"
    if d.startswith("PENNYMAC"):
        return "PENNYMAC"
    if d.startswith("WT FED"):
        return "WT FED"
    if d.startswith("EUNIFYPAY"):
        return "EUNIFYPAY"
    if d.startswith("ONLINE TRANSFER"):
        return "ONLINE TRANSFER"

    # normalized targets (defensive)
    if d.startswith("SHEGER MARKET"):
        return "SHEGER MARKET"
    if d.startswith("DOMINO'S PIZZA"):
        return "DOMINO'S PIZZA"
    if d.startswith("APPLEBEES"):
        return "APPLEBEES"
    if d.startswith("CHIPOTLE"):
        return "CHIPOTLE"
    if d.startswith("NAME-CHEAP.COM"):
        return "NAME-CHEAP.COM"
    if d.startswith("PRIMELENDING"):
        return "PRIMELENDING"

    tokens = d.split()
    if not tokens:
        return "OTHER"
    return " ".join(tokens[:2]) if len(tokens) >= 2 else tokens[0]


def group_key(description: str) -> str:
    """
    Default grouping:
    - ZELLE per person
    - else merchant_core(normalized_merchant)
    """
    d = normalize_merchant_name(description)
    if not d:
        return "OTHER"
    if d.startswith("ZELLE TO"):
        return f"ZELLE - {extract_zelle_person(d)}"
    return merchant_core(d)


def group_key_organized(description: str) -> str:
    """
    Organized grouping:
    - ALL ZELLE together
    - else merchant_core(normalized_merchant)
    """
    d = normalize_merchant_name(description)
    if not d:
        return "OTHER"
    if d.startswith("ZELLE TO"):
        return "ZELLE"
    return merchant_core(d)


def is_zelle_group(name: str) -> bool:
    return name.upper().startswith("ZELLE - ")


# -----------------------------
# Summaries + sorting
# -----------------------------
def sort_rows_for_detail(rows: List[Dict[str, Any]], key_fn: Callable[[str], str]) -> List[Dict[str, Any]]:
    rows.sort(
        key=lambda r: (
            key_fn(r.get("Description") or ""),
            (r.get("Description") or "").upper(),
            parse_date(r.get("Date")) or datetime.max,
        )
    )
    return rows


def build_summary(rows: List[Dict[str, Any]], key_fn: Callable[[str], str]) -> Dict[str, Dict[str, Any]]:
    summary: Dict[str, Dict[str, Any]] = {}
    for r in rows:
        g = key_fn(r.get("Description") or "")
        amt = parse_amount(r.get("Amount"))
        summary.setdefault(g, {"txns": 0, "total": 0.0})
        summary[g]["txns"] += 1
        summary[g]["total"] += amt
    return summary


def sort_summary_items(summary: Dict[str, Dict[str, Any]], sort_mode: str) -> List[Tuple[str, Dict[str, Any]]]:
    items = list(summary.items())
    if sort_mode == "total":
        return sorted(items, key=lambda kv: (-kv[1]["total"], -kv[1]["txns"], kv[0]))
    return sorted(items, key=lambda kv: (-kv[1]["txns"], kv[0], -kv[1]["total"]))


def apply_zelle_blocking(items_sorted: List[Tuple[str, Dict[str, Any]]], zelle_block: str):
    if zelle_block == "none":
        return items_sorted
    zelle_items = [kv for kv in items_sorted if is_zelle_group(kv[0])]
    other_items = [kv for kv in items_sorted if not is_zelle_group(kv[0])]
    return (zelle_items + other_items) if zelle_block == "first" else (other_items + zelle_items)


def reorder_priority_first(items_sorted: List[Tuple[str, Dict[str, Any]]], priority: List[str]) -> List[Tuple[str, Dict[str, Any]]]:
    lookup = {name: info for name, info in items_sorted}
    used = set()
    out: List[Tuple[str, Dict[str, Any]]] = []
    for p in priority:
        if p in lookup:
            out.append((p, lookup[p]))
            used.add(p)
    out.extend([(name, info) for name, info in items_sorted if name not in used])
    return out


# -----------------------------
# Excel reports
# -----------------------------
def write_excel_detail_grouped(headers, rows, xlsx_path: Path, key_fn):
    Workbook, Font = require_openpyxl()
    BOLD = Font(bold=True)

    ensure_required(headers, ["Description", "Amount"])
    amount_idx = headers.index("Amount") + 1
    desc_idx = headers.index("Description") + 1

    wb = Workbook()
    ws = wb.active
    ws.title = "Grouped Detail"

    ws.append([mt_timestamp_line("Generated (MT)")])
    ws.append(headers)

    ws.cell(row=1, column=1).font = BOLD
    for c in range(1, len(headers) + 1):
        ws.cell(row=2, column=c).font = BOLD

    def append_total(group_name, total_value, txn_count):
        row = [""] * len(headers)
        row[desc_idx - 1] = f"TOTAL ({group_name}) — {txn_count} txns"
        row[amount_idx - 1] = total_value
        ws.append(row)
        rr = ws.max_row
        ws.cell(row=rr, column=desc_idx).font = BOLD
        ws.cell(row=rr, column=amount_idx).font = BOLD
        ws.append([""] * len(headers))

    current_group = None
    group_total = 0.0
    group_count = 0

    for r in rows:
        g = key_fn(r.get("Description") or "")
        if current_group is not None and g != current_group:
            append_total(current_group, group_total, group_count)
            group_total = 0.0
            group_count = 0

        current_group = g
        group_total += parse_amount(r.get("Amount"))
        group_count += 1
        ws.append([r.get(h, "") for h in headers])

    if current_group is not None:
        append_total(current_group, group_total, group_count)

    for i in range(3, ws.max_row + 1):
        ws.cell(row=i, column=amount_idx).number_format = '"$"#,##0.00'

    wb.save(xlsx_path)


def write_excel_summary_items(items_sorted: List[Tuple[str, Dict[str, Any]]], xlsx_path: Path, title: str = "Family Summary"):
    Workbook, Font = require_openpyxl()
    BOLD = Font(bold=True)

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    ws.append([mt_timestamp_line("Generated (MT)")])
    ws.append(["Group", "Txns", "Total"])

    ws["A1"].font = BOLD
    ws["A2"].font = BOLD
    ws["B2"].font = BOLD
    ws["C2"].font = BOLD

    grand_txns = 0
    grand_total = 0.0

    for name, info in items_sorted:
        ws.append([name, info["txns"], info["total"]])
        grand_txns += info["txns"]
        grand_total += info["total"]

    ws.append(["GRAND TOTAL", grand_txns, grand_total])
    last = ws.max_row
    ws.cell(row=last, column=1).font = BOLD
    ws.cell(row=last, column=2).font = BOLD
    ws.cell(row=last, column=3).font = BOLD

    for r in range(3, ws.max_row + 1):
        ws.cell(row=r, column=3).number_format = '"$"#,##0.00'

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 16

    wb.save(xlsx_path)


def write_ready_to_print_excel(
    families_items: List[Tuple[str, Dict[str, Any]]],
    zelle_people_items: List[Tuple[str, Dict[str, Any]]],
    xlsx_path: Path,
):
    Workbook, Font = require_openpyxl()
    BOLD = Font(bold=True)

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Ready Summary"

    ws1.append([mt_timestamp_line("Generated (MT)")])
    ws1.append(["Families Summary (Ready to Print)"])
    ws1.append(["Group", "Txns", "Total"])

    ws1["A1"].font = BOLD
    ws1["A2"].font = BOLD
    ws1["A3"].font = BOLD
    ws1["B3"].font = BOLD
    ws1["C3"].font = BOLD

    gtx, gtot = 0, 0.0
    for name, info in families_items:
        ws1.append([name, info["txns"], info["total"]])
        gtx += info["txns"]
        gtot += info["total"]

    ws1.append(["GRAND TOTAL", gtx, gtot])
    last = ws1.max_row
    ws1[f"A{last}"].font = BOLD
    ws1[f"B{last}"].font = BOLD
    ws1[f"C{last}"].font = BOLD

    ws1.column_dimensions["A"].width = 42
    ws1.column_dimensions["B"].width = 10
    ws1.column_dimensions["C"].width = 16
    for r in range(4, ws1.max_row + 1):
        ws1.cell(row=r, column=3).number_format = '"$"#,##0.00'

    ws2 = wb.create_sheet("Zelle People")
    ws2.append([mt_timestamp_line("Generated (MT)")])
    ws2.append(["Zelle Transfers by Person"])
    ws2.append(["Group", "Txns", "Total"])

    ws2["A1"].font = BOLD
    ws2["A2"].font = BOLD
    ws2["A3"].font = BOLD
    ws2["B3"].font = BOLD
    ws2["C3"].font = BOLD

    ztx, ztot = 0, 0.0
    for name, info in zelle_people_items:
        ws2.append([name, info["txns"], info["total"]])
        ztx += info["txns"]
        ztot += info["total"]

    ws2.append(["GRAND TOTAL", ztx, ztot])
    last2 = ws2.max_row
    ws2[f"A{last2}"].font = BOLD
    ws2[f"B{last2}"].font = BOLD
    ws2[f"C{last2}"].font = BOLD

    ws2.column_dimensions["A"].width = 42
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 16
    for r in range(4, ws2.max_row + 1):
        ws2.cell(row=r, column=3).number_format = '"$"#,##0.00'

    wb.save(xlsx_path)


# -----------------------------
# PDF reports
# -----------------------------
def pdf_doc(pdf_path: Path, margin_in: float = 0.75):
    (letter, inch, colors, getSampleStyleSheet,
     SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak) = require_reportlab()
    doc = SimpleDocTemplate(
        str(pdf_path),
        pagesize=letter,
        leftMargin=margin_in * inch,
        rightMargin=margin_in * inch,
        topMargin=margin_in * inch,
        bottomMargin=margin_in * inch,
    )
    styles = getSampleStyleSheet()
    return (doc, styles, letter, inch, colors, Paragraph, Spacer, Table, TableStyle, PageBreak)


def style_summary_table(TableStyle, colors):
    return TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
    ])


def style_detail_table(TableStyle, colors):
    return TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.lightgrey),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ALIGN", (-1, 1), (-1, -1), "RIGHT"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
    ])


def write_pdf_quick_summary(
    items_sorted: List[Tuple[str, Dict[str, Any]]],
    pdf_path: Path,
    sort_mode: str,
    limit: int = 50,
    title_override: Optional[str] = None,
    removed_count: int = 0,
):
    """
    1-page summary PDF (table). Reusable for:
      - quick_pdf
      - exec_txns_desc
      - any future executive PDFs
    """
    doc, styles, _letter, inch, colors, Paragraph, Spacer, Table, TableStyle, _PageBreak = pdf_doc(pdf_path, margin_in=0.75)

    items_sorted = items_sorted[:max(0, int(limit))]
    title = title_override or (
        "Quick Summary — Sorted by Total (High → Low)" if sort_mode == "total"
        else "Quick Summary — Sorted by Transactions (High → Low)"
    )

    story = []
    story.append(Paragraph(title, styles["Title"]))
    story.append(Spacer(1, 0.08 * inch))
    story.append(Paragraph(mt_timestamp_line("Generated (MT)"), styles["Normal"]))

    if removed_count:
        story.append(Spacer(1, 0.06 * inch))
        story.append(Paragraph(f"Removed rows starting with '{REMOVE_DESC_PREFIX}': {removed_count}", styles["Normal"]))

    story.append(Spacer(1, 0.12 * inch))

    table_data = [["Group", "Txns", "Total"]]
    for name, info in items_sorted:
        table_data.append([name, str(info["txns"]), fmt_money(info["total"])])

    tbl = Table(table_data, colWidths=[3.6 * inch, 0.8 * inch, 1.4 * inch], repeatRows=1)
    tbl.setStyle(style_summary_table(TableStyle, colors))
    story.append(tbl)

    doc.build(story)


def write_pdf_summary(items_sorted, pdf_path: Path, title: str):
    doc, styles, _letter, inch, colors, Paragraph, Spacer, Table, TableStyle, _PageBreak = pdf_doc(pdf_path, margin_in=0.75)

    story = []
    story.append(Paragraph(title, styles["Title"]))
    story.append(Spacer(1, 0.08 * inch))
    story.append(Paragraph(mt_timestamp_line("Generated (MT)"), styles["Normal"]))
    story.append(Spacer(1, 0.18 * inch))

    table_data = [["Group", "Txns", "Total"]]
    gtx, gtot = 0, 0.0
    for name, info in items_sorted:
        table_data.append([name, str(info["txns"]), fmt_money(info["total"])])
        gtx += info["txns"]
        gtot += info["total"]
    table_data.append(["GRAND TOTAL", str(gtx), fmt_money(gtot)])

    tbl = Table(table_data, colWidths=[3.8 * inch, 0.8 * inch, 1.4 * inch], repeatRows=1)
    st = style_summary_table(TableStyle, colors)
    st.add("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold")
    st.add("BACKGROUND", (0, -1), (-1, -1), colors.whitesmoke)
    tbl.setStyle(st)

    story.append(tbl)
    doc.build(story)


def write_pdf_detail(rows, pdf_path: Path, key_fn: Callable[[str], str]):
    doc, styles, _letter, inch, colors, Paragraph, Spacer, Table, TableStyle, PageBreak = pdf_doc(pdf_path, margin_in=0.6)

    groups: Dict[str, List[Dict[str, Any]]] = {}
    for r in rows:
        g = key_fn(r.get("Description") or "")
        groups.setdefault(g, []).append(r)

    story = []
    story.append(Paragraph("Expenses — Detailed Grouped Report", styles["Title"]))
    story.append(Spacer(1, 0.08 * inch))
    story.append(Paragraph(mt_timestamp_line("Generated (MT)"), styles["Normal"]))
    story.append(Spacer(1, 0.18 * inch))

    for gname in sorted(groups.keys()):
        grows = groups[gname]
        grows.sort(key=lambda r: ((r.get("Description") or "").upper(), parse_date(r.get("Date")) or datetime.max))
        gtotal = sum(parse_amount(r.get("Amount")) for r in grows)

        story.append(Paragraph(
            f"<b>Group:</b> {gname} &nbsp;&nbsp; <b>Txns:</b> {len(grows)} &nbsp;&nbsp; <b>Total:</b> {fmt_money(gtotal)}",
            styles["Heading2"]
        ))
        story.append(Spacer(1, 0.08 * inch))

        table_data = [["Date", "Description", "Payee", "Payment Method", "Amount"]]
        for r in grows:
            table_data.append([
                (r.get("Date") or "").strip(),
                (r.get("Description") or "").strip(),
                (r.get("Payee") or "").strip(),
                (r.get("Payment Method") or "").strip(),
                fmt_money(parse_amount(r.get("Amount"))),
            ])

        tbl = Table(table_data,
                    colWidths=[0.9 * inch, 3.1 * inch, 1.4 * inch, 1.6 * inch, 0.9 * inch],
                    repeatRows=1)
        tbl.setStyle(style_detail_table(TableStyle, colors))
        story.append(tbl)
        story.append(PageBreak())

    doc.build(story)


def write_ready_to_print_pdf(
    families_items: List[Tuple[str, Dict[str, Any]]],
    zelle_people_items: List[Tuple[str, Dict[str, Any]]],
    pdf_path: Path,
):
    doc, styles, _letter, inch, colors, Paragraph, Spacer, Table, TableStyle, PageBreak = pdf_doc(pdf_path, margin_in=0.75)

    def table_from(items):
        data = [["Group", "Txns", "Total"]]
        gtx, gtot = 0, 0.0
        for name, info in items:
            data.append([name, str(info["txns"]), fmt_money(info["total"])])
            gtx += info["txns"]
            gtot += info["total"]
        data.append(["GRAND TOTAL", str(gtx), fmt_money(gtot)])
        tbl = Table(data, colWidths=[3.8 * inch, 0.8 * inch, 1.4 * inch], repeatRows=1)
        st = style_summary_table(TableStyle, colors)
        st.add("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold")
        st.add("BACKGROUND", (0, -1), (-1, -1), colors.whitesmoke)
        tbl.setStyle(st)
        return tbl

    story = []
    story.append(Paragraph("Ready to Print — Expense Summary", styles["Title"]))
    story.append(Spacer(1, 0.08 * inch))
    story.append(Paragraph(mt_timestamp_line("Generated (MT)"), styles["Normal"]))
    story.append(Spacer(1, 0.18 * inch))

    story.append(Paragraph("Families Summary", styles["Heading2"]))
    story.append(Spacer(1, 0.08 * inch))
    story.append(table_from(families_items))

    story.append(PageBreak())
    story.append(Paragraph("Zelle Transfers by Person", styles["Heading2"]))
    story.append(Spacer(1, 0.08 * inch))
    if zelle_people_items:
        story.append(table_from(zelle_people_items))
    else:
        story.append(Paragraph("No Zelle transfers found.", styles["Normal"]))

    doc.build(story)


# -----------------------------
# 18-month bucket PDF (one executive summary PDF with 4 windows)
# -----------------------------
def _in_range(d: datetime, start: datetime, end: datetime) -> bool:
    return start <= d <= end  # inclusive


def filter_rows_by_date_range(rows: List[Dict[str, Any]], start: datetime, end: datetime) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for r in rows:
        d = parse_date(r.get("Date"))
        if d and _in_range(d, start, end):
            out.append(r)
    return out


def write_pdf_quick_summary_18mo(
    rows: List[Dict[str, Any]],
    pdf_path: Path,
    buckets: List[Tuple[str, datetime, datetime]],
    sort_mode: str = "total",
    limit: int = 15,
    organized: bool = False,
):
    """
    Creates a single PDF page with 4 time buckets.
    Tip: keep --limit around 10–15 to ensure it fits on one page.
    """
    doc, styles, _letter, inch, colors, Paragraph, Spacer, Table, TableStyle, _PageBreak = pdf_doc(pdf_path, margin_in=0.55)

    key_fn = group_key_organized if organized else group_key

    story = []
    story.append(Paragraph("Quick Executive Summary — 18-Month Buckets", styles["Title"]))
    story.append(Spacer(1, 0.06 * inch))
    story.append(Paragraph(mt_timestamp_line("Generated (MT)"), styles["Normal"]))
    story.append(Spacer(1, 0.10 * inch))

    for (label, start, end) in buckets:
        bucket_rows = filter_rows_by_date_range(rows, start, end)

        story.append(Paragraph(f"<b>{label}</b>", styles["Heading3"]))
        story.append(Spacer(1, 0.02 * inch))

        if not bucket_rows:
            story.append(Paragraph("No transactions found in this range.", styles["Normal"]))
            story.append(Spacer(1, 0.10 * inch))
            continue

        summary = build_summary(bucket_rows, key_fn=key_fn)
        items = sort_summary_items(summary, sort_mode=sort_mode)[:max(0, int(limit))]

        bucket_txns = sum(info["txns"] for info in summary.values())
        bucket_total = sum(info["total"] for info in summary.values())

        story.append(Paragraph(f"Txns: <b>{bucket_txns}</b> &nbsp;&nbsp; Total: <b>{fmt_money(bucket_total)}</b>", styles["Normal"]))
        story.append(Spacer(1, 0.03 * inch))

        table_data = [["Group", "Txns", "Total"]]
        for name, info in items:
            table_data.append([name, str(info["txns"]), fmt_money(info["total"])])

        tbl = Table(table_data, colWidths=[3.15 * inch, 0.65 * inch, 1.25 * inch], repeatRows=1)
        tbl.setStyle(style_summary_table(TableStyle, colors))
        story.append(tbl)
        story.append(Spacer(1, 0.10 * inch))

    doc.build(story)


# -----------------------------
# Runners (commands)
# -----------------------------
def run_spacing_fix(in_path: Path, out_name: str):
    headers, rows = load_csv_rows(in_path)
    if not headers:
        raise ValueError("No headers found in CSV.")
    fixed = [{h: normalize_spaces(r.get(h, "")) for h in headers} for r in rows]
    out_csv = out_path("csv", out_name)
    write_csv_rows(out_csv, headers, fixed)
    print(mt_timestamp_line("Generated (MT)"))
    print(f"✅ Spacing fixed: {out_csv}")


def run_quick(in_path: Path, limit: int, sort_mode: str, organized: bool):
    _headers, rows = load_csv_rows(in_path)
    cleaned, _removed = clean_rows(rows)
    key_fn = group_key_organized if organized else group_key
    summary = build_summary(cleaned, key_fn=key_fn)
    items = sort_summary_items(summary, sort_mode=sort_mode)[:max(0, int(limit))]
    print(mt_timestamp_line("Generated (MT)"))
    print("✅ Quick Summary:")
    for name, info in items:
        print(f"  - {name}: {info['txns']} txns, {fmt_money(info['total'])}")


def run_quick_pdf(in_path: Path, out_pdf: str, limit: int, sort_mode: str, organized: bool):
    _headers, rows = load_csv_rows(in_path)
    cleaned, removed = clean_rows(rows)
    key_fn = group_key_organized if organized else group_key
    summary = build_summary(cleaned, key_fn=key_fn)
    items = sort_summary_items(summary, sort_mode=sort_mode)
    pdf_path = out_path("pdf", out_pdf)

    write_pdf_quick_summary(
        items_sorted=items,
        pdf_path=pdf_path,
        sort_mode=sort_mode,
        limit=limit,
        removed_count=removed,
    )

    print(mt_timestamp_line("Generated (MT)"))
    print(f"✅ Quick Summary PDF created: {pdf_path}")


def run_quick_pdf_18mo(in_path: Path, out_pdf: str, limit: int, sort_mode: str, organized: bool):
    _headers, rows = load_csv_rows(in_path)
    cleaned, _removed = clean_rows(rows)
    pdf_path = out_path("pdf", out_pdf)

    write_pdf_quick_summary_18mo(
        rows=cleaned,
        pdf_path=pdf_path,
        buckets=BUCKETS_18MO,
        sort_mode=sort_mode,
        limit=limit,
        organized=organized,
    )

    print(mt_timestamp_line("Generated (MT)"))
    print(f"✅ 18-month Executive Quick Summary PDF created: {pdf_path}")


def run_pipeline(
    in_path: Path,
    excel_detail_out: str,
    excel_summary_out: str,
    pdf_detail_out: str,
    pdf_summary_out: str,
    summary_sort: str,
):
    headers, rows = load_csv_rows(in_path)
    if not headers:
        raise ValueError("No headers found in CSV.")
    ensure_required(headers, ["Description", "Amount"])

    cleaned, _removed = clean_rows(rows)
    detail_rows = sort_rows_for_detail(cleaned, key_fn=group_key)
    summary = build_summary(detail_rows, key_fn=group_key)

    excel_detail_path = out_path("xlsx", excel_detail_out)
    excel_summary_path = out_path("xlsx", excel_summary_out)
    pdf_detail_path = out_path("pdf", pdf_detail_out)
    pdf_summary_path = out_path("pdf", pdf_summary_out)

    write_excel_detail_grouped(headers, detail_rows, excel_detail_path, key_fn=group_key)

    items = sort_summary_items(summary, sort_mode=summary_sort)
    write_excel_summary_items(items, excel_summary_path, title="Family Summary")

    write_pdf_detail(detail_rows, pdf_detail_path, key_fn=group_key)
    write_pdf_summary(items, pdf_summary_path, title="Expense Summary")

    print(mt_timestamp_line("Generated (MT)"))
    print("✅ Pipeline complete:")
    print(f"   - {excel_detail_path}")
    print(f"   - {excel_summary_path}")
    print(f"   - {pdf_detail_path}")
    print(f"   - {pdf_summary_path}")


def run_pdf_families(in_path: Path, out_pdf: str, zelle_block: str, sort_mode: str):
    _headers, rows = load_csv_rows(in_path)
    cleaned, _removed = clean_rows(rows)
    summary = build_summary(cleaned, key_fn=group_key)
    items = sort_summary_items(summary, sort_mode=sort_mode)
    items = apply_zelle_blocking(items, zelle_block=zelle_block)
    pdf_path = out_path("pdf", out_pdf)
    write_pdf_summary(items, pdf_path, title="Families Summary")
    print(mt_timestamp_line("Generated (MT)"))
    print(f"✅ PDF created: {pdf_path}")


def run_excel_families(in_path: Path, out_xlsx: str, zelle_block: str, sort_mode: str):
    _headers, rows = load_csv_rows(in_path)
    cleaned, _removed = clean_rows(rows)
    summary = build_summary(cleaned, key_fn=group_key)
    items = sort_summary_items(summary, sort_mode=sort_mode)
    items = apply_zelle_blocking(items, zelle_block=zelle_block)
    xlsx_path = out_path("xlsx", out_xlsx)
    write_excel_summary_items(items, xlsx_path, title="Family Summary Sorted")
    print(mt_timestamp_line("Generated (MT)"))
    print(f"✅ Excel created: {xlsx_path}")


def run_organized_pdf(in_path: Path, out_pdf: str, top_total: int):
    _headers, rows = load_csv_rows(in_path)
    cleaned, _removed = clean_rows(rows)
    summary = build_summary(cleaned, key_fn=group_key_organized)
    items_total = sort_summary_items(summary, sort_mode="total")[:max(0, int(top_total))]
    pdf_path = out_path("pdf", out_pdf)
    write_pdf_summary(items_total, pdf_path, title="Organized Report (Top by Total)")
    print(mt_timestamp_line("Generated (MT)"))
    print(f"✅ Organized PDF created: {pdf_path}")


def run_ready_to_print(in_path: Path, top_other: int):
    _headers, rows = load_csv_rows(in_path)
    cleaned, _removed = clean_rows(rows)

    # Families summary (ZELLE unified)
    families_summary = build_summary(cleaned, key_fn=group_key_organized)
    families_items = sort_summary_items(families_summary, sort_mode="total")
    families_items = reorder_priority_first(families_items, READY_FAMILIES_PRIORITY)

    # keep all priority + top N others
    if top_other is not None and top_other >= 0:
        priority_set = set(READY_FAMILIES_PRIORITY)
        kept_priority = [(n, i) for (n, i) in families_items if n in priority_set]
        others = [(n, i) for (n, i) in families_items if n not in priority_set]
        families_items = kept_priority + (others[:top_other] if top_other else [])

    # Zelle by person (ZELLE - Person)
    zelle_people_summary = build_summary(cleaned, key_fn=group_key)
    zelle_people_all = sort_summary_items(zelle_people_summary, sort_mode="total")
    zelle_people_items = [(n, i) for (n, i) in zelle_people_all if n.upper().startswith("ZELLE - ")]

    xlsx_path = out_path("xlsx", READY_TO_PRINT_XLSX)
    pdf_path = out_path("pdf", READY_TO_PRINT_PDF)
    write_ready_to_print_excel(families_items, zelle_people_items, xlsx_path)
    write_ready_to_print_pdf(families_items, zelle_people_items, pdf_path)

    print(mt_timestamp_line("Generated (MT)"))
    print("✅ Ready-to-print outputs created:")
    print(f"   - {xlsx_path}")
    print(f"   - {pdf_path}")


def run_exec_txns_desc(in_path: Path, out_pdf: str, limit: int, organized: bool):
    _headers, rows = load_csv_rows(in_path)
    cleaned, removed = clean_rows(rows)

    key_fn = group_key_organized if organized else group_key
    summary = build_summary(cleaned, key_fn=key_fn)

    items = sort_summary_items(summary, sort_mode="txns")

    pdf_path = out_path("pdf", out_pdf)
    write_pdf_quick_summary(
        items_sorted=items,
        pdf_path=pdf_path,
        sort_mode="txns",
        limit=limit,
        title_override="Quick Executive Summary — Highest to Lowest Transactions",
        removed_count=removed,
    )

    print(mt_timestamp_line("Generated (MT)"))
    print(f"✅ exec_txns_desc PDF created: {pdf_path}")


def run_all(in_path: Path):
    """
    Runs EVERYTHING in one shot:
      - pipeline
      - ready_to_print
      - quick_pdf
      - quick_pdf_18mo
      - exec_txns_desc
    """
    print(mt_timestamp_line("Generated (MT)"))
    print("🚀 Running ALL reports...")

    run_pipeline(
        in_path=in_path,
        excel_detail_out=DEFAULT_EXCEL_DETAIL_OUT,
        excel_summary_out=DEFAULT_EXCEL_SUMMARY_OUT,
        pdf_detail_out=DEFAULT_PDF_DETAIL_OUT,
        pdf_summary_out=DEFAULT_PDF_SUMMARY_OUT,
        summary_sort="txns",
    )

    run_ready_to_print(in_path, top_other=25)

    run_quick_pdf(
        in_path=in_path,
        out_pdf=DEFAULT_PDF_QUICK_OUT,
        limit=60,
        sort_mode="txns",
        organized=False,
    )

    run_quick_pdf_18mo(
        in_path=in_path,
        out_pdf=DEFAULT_PDF_QUICK_18MO_OUT,
        limit=15,
        sort_mode="total",
        organized=True,
    )

    run_exec_txns_desc(
        in_path=in_path,
        out_pdf=DEFAULT_PDF_HIGHEST_TXNS_OUT,
        limit=25,
        organized=True,
    )

    print("✅ ALL reports completed.")
    print("📂 Outputs created under:")
    print(f"   - {OUT_XLSX_DIR}")
    print(f"   - {OUT_PDF_DIR}")
    print(f"   - {OUT_CSV_DIR}")


# -----------------------------
# CLI
# -----------------------------
def main():
    p = argparse.ArgumentParser(description="Finance Master (single-file): clean + group + Excel/PDF outputs.")
    p.add_argument("--in", dest="input_csv", default=DEFAULT_INPUT_CSV, help="Input CSV filename (same folder).")

    sub = p.add_subparsers(dest="cmd", required=True)

    s = sub.add_parser("spacing", help="Fix inconsistent spacing in raw CSV (no grouping, no deletions).")
    s.add_argument("--out", default=DEFAULT_SPACING_OUT, help="Output CSV filename.")

    q = sub.add_parser("quick", help="Print quick summary to console.")
    q.add_argument("--limit", type=int, default=50)
    q.add_argument("--sort", choices=["txns", "total"], default="txns")
    q.add_argument("--organized", action="store_true", help="Use organized grouping (ALL ZELLE together).")

    qp = sub.add_parser("quick_pdf", help="Create a 1-page Quick Summary PDF.")
    qp.add_argument("--out", default=DEFAULT_PDF_QUICK_OUT)
    qp.add_argument("--limit", type=int, default=60)
    qp.add_argument("--sort", choices=["txns", "total"], default="txns")
    qp.add_argument("--organized", action="store_true", help="Use organized grouping (ALL ZELLE together).")

    ex = sub.add_parser(
        "exec_txns_desc",
        help="Create a 1-page executive summary sorted by transaction count (highest → lowest)."
    )
    ex.add_argument("--out", default=DEFAULT_PDF_HIGHEST_TXNS_OUT)
    ex.add_argument("--limit", type=int, default=25)
    ex.add_argument("--organized", action="store_true", help="Use organized grouping (ALL ZELLE together).")

    q18 = sub.add_parser("quick_pdf_18mo", help="Create a 1-page executive summary PDF split into 18-month buckets.")
    q18.add_argument("--out", default=DEFAULT_PDF_QUICK_18MO_OUT)
    q18.add_argument("--limit", type=int, default=15)
    q18.add_argument("--sort", choices=["txns", "total"], default="total")
    q18.add_argument("--organized", action="store_true", help="Use organized grouping (ALL ZELLE together).")

    pl = sub.add_parser("pipeline", help="Excel detail+summary + PDF detail+summary.")
    pl.add_argument("--excel-detail-out", default=DEFAULT_EXCEL_DETAIL_OUT)
    pl.add_argument("--excel-summary-out", default=DEFAULT_EXCEL_SUMMARY_OUT)
    pl.add_argument("--pdf-detail-out", default=DEFAULT_PDF_DETAIL_OUT)
    pl.add_argument("--pdf-summary-out", default=DEFAULT_PDF_SUMMARY_OUT)
    pl.add_argument("--summary-sort", choices=["txns", "total"], default="txns")

    sub.add_parser("all", help="Run EVERYTHING at once: pipeline + ready_to_print + quick_pdf + quick_pdf_18mo + exec_txns_desc.")

    pf = sub.add_parser("pdf_families", help="PDF families summary (sorted).")
    pf.add_argument("--out", default=DEFAULT_PDF_FAMILIES_SORTED_OUT)
    pf.add_argument("--zelle-block", choices=["first", "last", "none"], default="first")
    pf.add_argument("--sort", choices=["total", "txns"], default="total")

    ef = sub.add_parser("excel_families", help="Excel families summary (sorted).")
    ef.add_argument("--out", default=DEFAULT_EXCEL_FAMILIES_OUT)
    ef.add_argument("--zelle-block", choices=["first", "last", "none"], default="first")
    ef.add_argument("--sort", choices=["total", "txns"], default="total")

    op = sub.add_parser("organized_pdf", help="Organized PDF (Top by Total).")
    op.add_argument("--out", default=DEFAULT_PDF_ORGANIZED_OUT)
    op.add_argument("--top-total", type=int, default=25)

    rtp = sub.add_parser("ready_to_print", help="Create ready_to_print.xlsx and ready_to_print.pdf.")
    rtp.add_argument("--top-other", type=int, default=25)

    args = p.parse_args()

    base = Path(__file__).parent
    in_path = base / args.input_csv
    if not in_path.exists():
        raise FileNotFoundError(f"Input CSV not found: {in_path}")

    if args.cmd == "spacing":
        run_spacing_fix(in_path, args.out)

    elif args.cmd == "quick":
        run_quick(in_path, limit=args.limit, sort_mode=args.sort, organized=args.organized)

    elif args.cmd == "quick_pdf":
        run_quick_pdf(in_path, out_pdf=args.out, limit=args.limit, sort_mode=args.sort, organized=args.organized)

    elif args.cmd == "exec_txns_desc":
        run_exec_txns_desc(in_path, out_pdf=args.out, limit=args.limit, organized=args.organized)

    elif args.cmd == "quick_pdf_18mo":
        run_quick_pdf_18mo(in_path, out_pdf=args.out, limit=args.limit, sort_mode=args.sort, organized=args.organized)

    elif args.cmd == "pipeline":
        run_pipeline(
            in_path=in_path,
            excel_detail_out=args.excel_detail_out,
            excel_summary_out=args.excel_summary_out,
            pdf_detail_out=args.pdf_detail_out,
            pdf_summary_out=args.pdf_summary_out,
            summary_sort=args.summary_sort,
        )

    elif args.cmd == "pdf_families":
        run_pdf_families(in_path, out_pdf=args.out, zelle_block=args.zelle_block, sort_mode=args.sort)

    elif args.cmd == "excel_families":
        run_excel_families(in_path, out_xlsx=args.out, zelle_block=args.zelle_block, sort_mode=args.sort)

    elif args.cmd == "organized_pdf":
        run_organized_pdf(in_path, out_pdf=args.out, top_total=args.top_total)

    elif args.cmd == "ready_to_print":
        run_ready_to_print(in_path, top_other=args.top_other)

    elif args.cmd == "all":
        run_all(in_path)


if __name__ == "__main__":
    main()
