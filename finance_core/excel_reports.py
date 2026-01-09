"""
finance_core.excel_reports
Excel creation (openpyxl).
"""
from __future__ import annotations
from pathlib import Path
from typing import Any, Callable, Dict, List, Tuple
from .io_csv import ensure_required
from .parsing import parse_amount
from .utils import mt_timestamp_line

def require_openpyxl():
    try:
        from openpyxl import Workbook  # noqa
        from openpyxl.styles import Font  # noqa
        return Workbook, Font
    except Exception:
        raise SystemExit("Missing dependency: openpyxl\nInstall with: pip3 install openpyxl\n")

def write_excel_detail_grouped(headers: List[str], rows: List[Dict[str, Any]], xlsx_path: Path, key_fn: Callable[[str], str]) -> None:
    Workbook, Font = require_openpyxl()
    BOLD = Font(bold=True)

    ensure_required(headers, ["Description", "Amount"])
    amount_idx = headers.index("Amount") + 1
    desc_idx = headers.index("Description") + 1

    wb = Workbook()
    ws = wb.active
    ws.title = "Grouped Detail"

    ws.append([mt_timestamp_line("Generated (MT)")])
    ws.append(headers)

    ws.cell(row=1, column=1).font = BOLD
    for c in range(1, len(headers) + 1):
        ws.cell(row=2, column=c).font = BOLD

    def append_total(group_name: str, total_value: float, txn_count: int):
        row = [""] * len(headers)
        row[desc_idx - 1] = f"TOTAL ({group_name}) — {txn_count} txns"
        row[amount_idx - 1] = total_value
        ws.append(row)
        rr = ws.max_row
        ws.cell(row=rr, column=desc_idx).font = BOLD
        ws.cell(row=rr, column=amount_idx).font = BOLD
        ws.append([""] * len(headers))

    current_group = None
    group_total = 0.0
    group_count = 0

    for r in rows:
        g = key_fn(r.get("Description") or "")
        if current_group is not None and g != current_group:
            append_total(current_group, group_total, group_count)
            group_total = 0.0
            group_count = 0

        current_group = g
        group_total += parse_amount(r.get("Amount"))
        group_count += 1
        ws.append([r.get(h, "") for h in headers])

    if current_group is not None:
        append_total(current_group, group_total, group_count)

    for i in range(3, ws.max_row + 1):
        ws.cell(row=i, column=amount_idx).number_format = '"$"#,##0.00'

    wb.save(xlsx_path)

def write_excel_summary_items(items_sorted: List[Tuple[str, Dict[str, Any]]], xlsx_path: Path, title: str = "Family Summary") -> None:
    Workbook, Font = require_openpyxl()
    BOLD = Font(bold=True)

    wb = Workbook()
    ws = wb.active
    ws.title = title[:31]

    ws.append([mt_timestamp_line("Generated (MT)")])
    ws.append(["Group", "Txns", "Total"])

    ws["A1"].font = BOLD
    ws["A2"].font = BOLD
    ws["B2"].font = BOLD
    ws["C2"].font = BOLD

    grand_txns = 0
    grand_total = 0.0

    for name, info in items_sorted:
        ws.append([name, info["txns"], info["total"]])
        grand_txns += info["txns"]
        grand_total += info["total"]

    ws.append(["GRAND TOTAL", grand_txns, grand_total])
    last = ws.max_row
    ws.cell(row=last, column=1).font = BOLD
    ws.cell(row=last, column=2).font = BOLD
    ws.cell(row=last, column=3).font = BOLD

    for r in range(3, ws.max_row + 1):
        ws.cell(row=r, column=3).number_format = '"$"#,##0.00'

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 16

    wb.save(xlsx_path)

def write_ready_to_print_excel(
    families_items: List[Tuple[str, Dict[str, Any]]],
    zelle_people_items: List[Tuple[str, Dict[str, Any]]],
    xlsx_path: Path,
) -> None:
    Workbook, Font = require_openpyxl()
    BOLD = Font(bold=True)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Ready Summary"
    ws1.append([mt_timestamp_line("Generated (MT)")])
    ws1.append(["Families Summary (Ready to Print)"])
    ws1.append(["Group", "Txns", "Total"])

    for cell in ("A1","A2","A3","B3","C3"):
        ws1[cell].font = BOLD

    gtx, gtot = 0, 0.0
    for name, info in families_items:
        ws1.append([name, info["txns"], info["total"]])
        gtx += info["txns"]
        gtot += info["total"]
    ws1.append(["GRAND TOTAL", gtx, gtot])
    last = ws1.max_row
    for c in (1,2,3):
        ws1.cell(row=last, column=c).font = BOLD

    ws1.column_dimensions["A"].width = 42
    ws1.column_dimensions["B"].width = 10
    ws1.column_dimensions["C"].width = 16
    for r in range(4, ws1.max_row + 1):
        ws1.cell(row=r, column=3).number_format = '"$"#,##0.00'

    ws2 = wb.create_sheet("Zelle People")
    ws2.append([mt_timestamp_line("Generated (MT)")])
    ws2.append(["Zelle Transfers by Person"])
    ws2.append(["Group", "Txns", "Total"])
    for cell in ("A1","A2","A3","B3","C3"):
        ws2[cell].font = BOLD

    ztx, ztot = 0, 0.0
    for name, info in zelle_people_items:
        ws2.append([name, info["txns"], info["total"]])
        ztx += info["txns"]
        ztot += info["total"]
    ws2.append(["GRAND TOTAL", ztx, ztot])
    last2 = ws2.max_row
    for c in (1,2,3):
        ws2.cell(row=last2, column=c).font = BOLD

    ws2.column_dimensions["A"].width = 42
    ws2.column_dimensions["B"].width = 10
    ws2.column_dimensions["C"].width = 16
    for r in range(4, ws2.max_row + 1):
        ws2.cell(row=r, column=3).number_format = '"$"#,##0.00'

    wb.save(xlsx_path)
