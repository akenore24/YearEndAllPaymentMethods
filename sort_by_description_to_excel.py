#!/usr/bin/env python3
"""
Finance CSV -> Clean + Normalize + Group Families + Totals (Excel)

- Normalizes Description spacing
- Removes rows where Description starts with "ONLINE TRANSFER REF"
- Normalizes Payment Method (WF Active Cash Visa -> WFACV...####)
- Groups transactions into merchant "families" (ZELLE, AMAZON, etc.)
- Sorts by Family -> Description -> Date
- Writes Excel with TOTAL row + blank separator after each family
"""

import csv
import re
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font


# -----------------------------
# Config
# -----------------------------
INPUT_CSV = "expenses.csv"
OUTPUT_XLSX = "expenses_grouped_families_totals.xlsx"

REMOVE_DESC_PREFIX = "ONLINE TRANSFER REF"

WF_CARD_PREFIX = "WELLS FARGO ACTIVE CASH VISA(R) CARD"
WF_CARD_ALIAS = "WFACV"

DATE_FORMATS = (
    "%m/%d/%Y",
    "%m/%d/%y",
    "%Y-%m-%d",
    "%m-%d-%Y",
    "%m-%d-%y",
)

BOLD = Font(bold=True)


# -----------------------------
# Helpers
# -----------------------------
def normalize_spaces(text: str) -> str:
    return " ".join((text or "").split()).strip()


def normalize_payment_method(value: str) -> str:
    value = (value or "").strip()
    if value.upper().startswith(WF_CARD_PREFIX):
        suffix = value[len(WF_CARD_PREFIX):].strip()
        return f"{WF_CARD_ALIAS}{suffix}"
    return value


def parse_date(value: str):
    s = ("" if value is None else str(value)).strip()
    if not s:
        return None

    s = s.split()[0]
    if "T" in s:
        s = s.split("T")[0]

    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def parse_amount(value) -> float:
    if value is None:
        return 0.0
    s = str(value).strip()
    if not s:
        return 0.0

    negative = False
    if s.startswith("(") and s.endswith(")"):
        negative = True
        s = s[1:-1].strip()

    s = s.replace("$", "").replace(",", "")
    try:
        n = float(s)
        return -n if negative else n
    except ValueError:
        return 0.0


def family_key(description: str) -> str:
    """
    Group similar descriptions into vendor families.

    You asked specifically:
      - All Zelle -> one group
      - All Amazon -> one group
      - Follow same for others

    This function is where you expand rules safely.
    """
    d = normalize_spaces(description).upper()

    if not d:
        return "OTHER"

    # ---- Explicit "families" ----
    if d.startswith("ZELLE TO"):
        return "ZELLE"

    if d.startswith("AMAZON"):
        return "AMAZON"

    if d.startswith("APPLE.COM/BILL") or d.startswith("APPLE"):
        return "APPLE"

    if d.startswith("ATM WITHDRAWAL"):
        return "ATM WITHDRAWAL"

    if d.startswith("COMCAST") or d.startswith("XFINITY"):
        return "COMCAST/XFINITY"

    if d.startswith("COSTCO GAS"):
        return "COSTCO GAS"
    if d.startswith("COSTCO WHSE") or d.startswith("COSTCO WHOLESALE"):
        return "COSTCO WHSE"

    if d.startswith("WAL-MART") or d.startswith("WM SUPERCENTER"):
        return "WALMART"

    if d.startswith("KING SOOPERS"):
        return "KING SOOPERS"

    if d.startswith("SPROUTS"):
        return "SPROUTS"

    if d.startswith("WHOLEFDS") or d.startswith("WHOLE FOODS"):
        return "WHOLE FOODS"

    if d.startswith("STATE FARM"):
        return "STATE FARM"

    if d.startswith("DEPT EDUCATION") or "STUDENT LN" in d:
        return "STUDENT LOAN"

    if d.startswith("PENNYMAC"):
        return "PENNYMAC"

    if d.startswith("E 470") or "EXPRESS TOLLS" in d:
        return "TOLLS"

    # ---- Generic fallback (still useful) ----
    # Use first token (or first 2 tokens) as a reasonable grouping
    tokens = d.split()
    if not tokens:
        return "OTHER"

    if len(tokens) >= 2 and (tokens[1].isdigit() or re.fullmatch(r"#\d+", tokens[1] or "")):
        return tokens[0]
    return " ".join(tokens[:2]) if len(tokens) >= 2 else tokens[0]


# -----------------------------
# Load / clean / sort
# -----------------------------
def load_rows(csv_path: Path):
    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        return reader.fieldnames, list(reader)


def clean_rows(rows):
    removed = 0
    cleaned = []

    for row in rows:
        row["Description"] = normalize_spaces(row.get("Description"))

        desc_upper = (row.get("Description") or "").upper()
        if desc_upper.startswith(REMOVE_DESC_PREFIX.upper()):
            removed += 1
            continue

        row["Payment Method"] = normalize_payment_method(row.get("Payment Method"))
        cleaned.append(row)

    return cleaned, removed


def sort_rows(rows):
    rows.sort(
        key=lambda r: (
            family_key(r.get("Description") or ""),
            (r.get("Description") or "").upper(),
            parse_date(r.get("Date")) or datetime.max,
        )
    )
    return rows


# -----------------------------
# Write grouped Excel
# -----------------------------
def write_grouped_xlsx(headers, rows, xlsx_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Grouped Families"

    if "Amount" not in headers or "Description" not in headers:
        raise ValueError("CSV must contain 'Description' and 'Amount' columns.")

    amount_idx = headers.index("Amount") + 1
    desc_idx = headers.index("Description") + 1

    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = BOLD

    def append_total_row(group_name, total_value, txn_count):
        row = [""] * len(headers)
        row[desc_idx - 1] = f"TOTAL ({group_name}) — {txn_count} txns"
        row[amount_idx - 1] = total_value
        ws.append(row)

        r = ws.max_row
        ws.cell(row=r, column=desc_idx).font = BOLD
        ws.cell(row=r, column=amount_idx).font = BOLD

        ws.append([""] * len(headers))  # blank separator

    current_group = None
    group_total = 0.0
    group_count = 0

    for row in rows:
        g = family_key(row.get("Description") or "")

        if current_group is not None and g != current_group:
            append_total_row(current_group, group_total, group_count)
            group_total = 0.0
            group_count = 0

        current_group = g
        group_total += parse_amount(row.get("Amount"))
        group_count += 1

        ws.append([row.get(h, "") for h in headers])

    if current_group is not None:
        append_total_row(current_group, group_total, group_count)

    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=amount_idx).number_format = "#,##0.00"

    wb.save(xlsx_path)


# -----------------------------
# Main
# -----------------------------
def main():
    base_dir = Path(__file__).parent
    csv_path = base_dir / INPUT_CSV
    xlsx_path = base_dir / OUTPUT_XLSX

    if not csv_path.exists():
        raise FileNotFoundError(f"Input CSV not found: {csv_path}")

    headers, rows = load_rows(csv_path)
    cleaned, removed = clean_rows(rows)
    sorted_rows = sort_rows(cleaned)
    write_grouped_xlsx(headers, sorted_rows, xlsx_path)

    print("✅ Done")
    print(f"🧹 Removed rows (prefix '{REMOVE_DESC_PREFIX}'): {removed}")
    print("🧾 Grouped by families (ZELLE, AMAZON, etc.) and added TOTAL rows.")
    print(f"📁 Output: {xlsx_path.name}")


if __name__ == "__main__":
    main()
#!/usr/bin/env python3
"""
Finance CSV -> Clean + Normalize + Sort + Group Totals (Excel)

- Normalizes Description spacing
- Removes "ONLINE TRANSFER REF..." rows
- Normalizes Payment Method (WF Active Cash Visa -> WFACV...####)
- Sorts by Group Key (merchant), then full Description, then Date
- Writes Excel with a TOTAL row + blank separator after each group
"""

import csv
import re
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font


# -----------------------------
# Config
# -----------------------------
INPUT_CSV = "expenses.csv"
OUTPUT_XLSX = "expenses_grouped_totals.xlsx"

REMOVE_DESC_PREFIX = "ONLINE TRANSFER REF"

WF_CARD_PREFIX = "WELLS FARGO ACTIVE CASH VISA(R) CARD"
WF_CARD_ALIAS = "WFACV"

DATE_FORMATS = (
    "%m/%d/%Y",
    "%m/%d/%y",
    "%Y-%m-%d",
    "%m-%d-%Y",
    "%m-%d-%y",
)

BOLD = Font(bold=True)


# -----------------------------
# Helpers
# -----------------------------
def normalize_spaces(text: str) -> str:
    return " ".join((text or "").split()).strip()


def normalize_payment_method(value: str) -> str:
    value = (value or "").strip()
    if value.upper().startswith(WF_CARD_PREFIX):
        suffix = value[len(WF_CARD_PREFIX):].strip()  # keeps "...4321"
        return f"{WF_CARD_ALIAS}{suffix}"
    return value


def parse_date(value: str):
    s = ("" if value is None else str(value)).strip()
    if not s:
        return None

    s = s.split()[0]
    if "T" in s:
        s = s.split("T")[0]

    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def parse_amount(value) -> float:
    if value is None:
        return 0.0
    s = str(value).strip()
    if not s:
        return 0.0

    negative = False
    if s.startswith("(") and s.endswith(")"):
        negative = True
        s = s[1:-1].strip()

    s = s.replace("$", "").replace(",", "")
    try:
        n = float(s)
        return -n if negative else n
    except ValueError:
        return 0.0


def merchant_key(description: str) -> str:
    """
    Extract a "similar name" key from Description.

    Rules (practical for bank exports):
    - Trim/collapse spaces
    - Remove trailing location chunks like "AURORA ,CO" from keying logic by focusing on leading merchant name
    - If it starts with patterns like:
        "7-ELEVEN 21494 ..." -> "7-ELEVEN"
        "CHIPOTLE 0871 ..." -> "CHIPOTLE"
        "DOMINO'S 6299 ..." -> "DOMINO'S"
    - If it starts with "COSTCO GAS #1652" -> "COSTCO GAS"
    - If it starts with "COSTCO WHSE #1652" -> "COSTCO WHSE"
    - If it starts with "AMAZON MKTPL*XXXX" -> "AMAZON MKTPL"
    - Otherwise: take first 2 tokens unless second token looks like a store number/#code, then first token.
    """
    d = normalize_spaces(description).upper()

    if not d:
        return ""

    # Common explicit patterns
    if d.startswith("AMAZON MKTPL"):
        return "AMAZON MKTPL"
    if d.startswith("APPLE.COM/BILL"):
        return "APPLE.COM/BILL"
    if d.startswith("COSTCO GAS"):
        return "COSTCO GAS"
    if d.startswith("COSTCO WHSE"):
        return "COSTCO WHSE"
    if d.startswith("ATM WITHDRAWAL"):
        return "ATM WITHDRAWAL"
    if d.startswith("PURCHASE AUTHORIZED ON"):
        return "PURCHASE AUTHORIZED ON"
    if d.startswith("ZELLE TO"):
        return "ZELLE TO"

    tokens = d.split()

    # If second token is numeric (store id) or looks like "#1234", group by first token
    if len(tokens) >= 2 and (tokens[1].isdigit() or re.fullmatch(r"#\d+", tokens[1] or "")):
        return tokens[0]

    # If first token has a hyphen merchant like "7-ELEVEN", that is usually enough
    # Otherwise take first 2 tokens for nicer grouping ("STATE FARM", "KING SOOPERS")
    if len(tokens) >= 2:
        return f"{tokens[0]} {tokens[1]}"
    return tokens[0]


# -----------------------------
# Load / clean
# -----------------------------
def load_rows(csv_path: Path):
    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        return reader.fieldnames, list(reader)


def clean_rows(rows):
    removed = 0
    cleaned = []

    for row in rows:
        row["Description"] = normalize_spaces(row.get("Description"))
        if (row["Description"] or "").upper().startswith(REMOVE_DESC_PREFIX.upper()):
            removed += 1
            continue

        row["Payment Method"] = normalize_payment_method(row.get("Payment Method"))
        cleaned.append(row)

    return cleaned, removed


def sort_rows(rows):
    rows.sort(
        key=lambda r: (
            merchant_key(r.get("Description") or ""),
            (r.get("Description") or "").upper(),
            parse_date(r.get("Date")) or datetime.max,
        )
    )
    return rows


# -----------------------------
# Write grouped Excel
# -----------------------------
def write_grouped_xlsx(headers, rows, xlsx_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Grouped Totals"

    if "Amount" not in headers or "Description" not in headers:
        raise ValueError("CSV must contain 'Description' and 'Amount' columns.")

    amount_idx = headers.index("Amount") + 1
    desc_idx = headers.index("Description") + 1

    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = BOLD

    def append_total_row(group_name, total_value, txn_count):
        total_row = [""] * len(headers)
        total_row[desc_idx - 1] = f"TOTAL ({group_name}) — {txn_count} txns"
        total_row[amount_idx - 1] = total_value
        ws.append(total_row)

        r = ws.max_row
        ws.cell(row=r, column=desc_idx).font = BOLD
        ws.cell(row=r, column=amount_idx).font = BOLD

        # blank separator row
        ws.append([""] * len(headers))

    current_group = None
    group_total = 0.0
    group_count = 0

    for row in rows:
        g = merchant_key(row.get("Description") or "")

        # group boundary
        if current_group is not None and g != current_group:
            append_total_row(current_group, group_total, group_count)
            group_total = 0.0
            group_count = 0

        current_group = g
        group_total += parse_amount(row.get("Amount"))
        group_count += 1

        ws.append([row.get(h, "") for h in headers])

    # final group
    if current_group is not None:
        append_total_row(current_group, group_total, group_count)

    # format Amount column
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=amount_idx).number_format = "#,##0.00"

    wb.save(xlsx_path)


# -----------------------------
# Main
# -----------------------------
def main():
    base_dir = Path(__file__).parent
    csv_path = base_dir / INPUT_CSV
    xlsx_path = base_dir / OUTPUT_XLSX

    if not csv_path.exists():
        raise FileNotFoundError(f"Input CSV not found: {csv_path}")

    headers, rows = load_rows(csv_path)
    cleaned, removed = clean_rows(rows)
    sorted_rows = sort_rows(cleaned)
    write_grouped_xlsx(headers, sorted_rows, xlsx_path)

    print("✅ Done")
    print(f"🧹 Removed rows (prefix '{REMOVE_DESC_PREFIX}'): {removed}")
    print("🧾 Grouped by 'similar merchant name' and added TOTAL rows.")
    print(f"📁 Output: {xlsx_path.name}")


if __name__ == "__main__":
    main()
#!/usr/bin/env python3
"""
Finance CSV Cleaner -> Grouped Excel with Subtotals

Does:
1) Normalize Description spacing
2) Remove rows where Description starts with "ONLINE TRANSFER REF"
3) Normalize Payment Method (WF Active Cash Visa -> WFACV...####)
4) Sort by Description (A→Z), then Date (oldest→newest)
5) Write Excel with:
   - Each Description group together
   - A TOTAL row after each group
   - A blank separator row after each TOTAL
"""

import csv
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font


# -----------------------------
# Config (edit if needed)
# -----------------------------
INPUT_CSV = "expenses.csv"
OUTPUT_XLSX = "expenses_grouped_with_totals.xlsx"

REMOVE_DESC_PREFIX = "ONLINE TRANSFER REF"

WF_CARD_PREFIX = "WELLS FARGO ACTIVE CASH VISA(R) CARD"
WF_CARD_ALIAS = "WFACV"

DATE_FORMATS = (
    "%m/%d/%Y",
    "%m/%d/%y",
    "%Y-%m-%d",
    "%m-%d-%Y",
    "%m-%d-%y",
)


# -----------------------------
# Helpers
# -----------------------------
def normalize_spaces(text: str) -> str:
    return " ".join((text or "").split()).strip()


def normalize_payment_method(value: str) -> str:
    value = (value or "").strip()
    if value.upper().startswith(WF_CARD_PREFIX):
        suffix = value[len(WF_CARD_PREFIX):].strip()  # keeps "...4321"
        return f"{WF_CARD_ALIAS}{suffix}"
    return value


def parse_date(value: str):
    s = ("" if value is None else str(value)).strip()
    if not s:
        return None

    s = s.split()[0]
    if "T" in s:
        s = s.split("T")[0]

    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def parse_amount(value) -> float:
    """
    Convert Amount cell to float safely.
    Handles: "$1,234.56", "1234.56", "(123.45)"
    """
    if value is None:
        return 0.0

    s = str(value).strip()
    if not s:
        return 0.0

    negative = False
    if s.startswith("(") and s.endswith(")"):
        negative = True
        s = s[1:-1].strip()

    s = s.replace("$", "").replace(",", "")
    try:
        n = float(s)
        return -n if negative else n
    except ValueError:
        return 0.0


# -----------------------------
# Load / clean / sort
# -----------------------------
def load_rows(csv_path: Path):
    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        return reader.fieldnames, list(reader)


def clean_rows(rows):
    removed = 0
    cleaned = []

    for row in rows:
        row["Description"] = normalize_spaces(row.get("Description"))

        desc_key = (row.get("Description") or "").upper()
        if desc_key.startswith(REMOVE_DESC_PREFIX.upper()):
            removed += 1
            continue

        row["Payment Method"] = normalize_payment_method(row.get("Payment Method"))
        cleaned.append(row)

    return cleaned, removed


def sort_rows(rows):
    rows.sort(
        key=lambda r: (
            (r.get("Description") or "").upper(),
            parse_date(r.get("Date")) or datetime.max,
        )
    )
    return rows


# -----------------------------
# Excel writer: grouped + totals
# -----------------------------
def write_grouped_xlsx(headers, rows, xlsx_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Grouped Totals"

    # Find column indexes
    try:
        amount_col_idx = headers.index("Amount") + 1  # 1-based for Excel
    except ValueError:
        raise ValueError("Could not find 'Amount' column in headers.")

    try:
        desc_col_idx = headers.index("Description") + 1
    except ValueError:
        raise ValueError("Could not find 'Description' column in headers.")

    # Header row
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        ws.cell(row=1, column=c).font = Font(bold=True)

    bold = Font(bold=True)

    current_desc = None
    group_total = 0.0
    group_count = 0

    def write_total_and_separator(desc, total_value, count_value):
        """Write subtotal row + blank row."""
        if desc is None:
            return

        total_row = [""] * len(headers)
        total_row[desc_col_idx - 1] = f"TOTAL ({desc}) — {count_value} txns"
        total_row[amount_col_idx - 1] = total_value

        ws.append(total_row)

        # Bold the subtotal line
        r = ws.max_row
        ws.cell(row=r, column=desc_col_idx).font = bold
        ws.cell(row=r, column=amount_col_idx).font = bold

        # Blank separator row
        ws.append([""] * len(headers))

    for row in rows:
        desc = row.get("Description") or ""

        # Group boundary: if Description changes, write subtotal for previous group
        if current_desc is not None and desc != current_desc:
            write_total_and_separator(current_desc, group_total, group_count)
            group_total = 0.0
            group_count = 0

        current_desc = desc
        group_count += 1
        group_total += parse_amount(row.get("Amount"))

        # Write the actual row
        ws.append([row.get(h, "") for h in headers])

    # Final group subtotal
    write_total_and_separator(current_desc, group_total, group_count)

    # Optional: number format for Amount column (Excel)
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=amount_col_idx).number_format = "#,##0.00"

    wb.save(xlsx_path)


# -----------------------------
# Main
# -----------------------------
def main():
    base_dir = Path(__file__).parent
    csv_path = base_dir / INPUT_CSV
    xlsx_path = base_dir / OUTPUT_XLSX

    if not csv_path.exists():
        raise FileNotFoundError(f"Input CSV not found: {csv_path}")

    headers, rows = load_rows(csv_path)
    cleaned, removed = clean_rows(rows)
    sorted_rows = sort_rows(cleaned)

    write_grouped_xlsx(headers, sorted_rows, xlsx_path)

    print("✅ Done")
    print(f"🧹 Removed rows (prefix '{REMOVE_DESC_PREFIX}'): {removed}")
    print("📌 Output includes a TOTAL row + blank separator after each Description group.")
    print(f"📁 Output: {xlsx_path.name}")


if __name__ == "__main__":
    main()
#!/usr/bin/env python3
"""
Finance CSV Cleaner -> Excel

What it does (DRY + organized):
1) Normalize Description spacing (collapse multiple spaces, trim).
2) Remove rows where Description starts with "ONLINE TRANSFER REF" (case-insensitive).
3) Normalize Payment Method:
   "WELLS FARGO ACTIVE CASH VISA(R) CARD ...4321" -> "WFACV...4321"
4) Sort by Description (A→Z), then Date (oldest→newest).
5) Export to a single .xlsx file in the same folder as this script.
"""

import csv
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook


# -----------------------------
# Config (edit these if needed)
# -----------------------------
INPUT_CSV = "expenses.csv"
OUTPUT_XLSX = "expenses_clean_sorted.xlsx"
REMOVE_DESC_PREFIX = "ONLINE TRANSFER REF"

WF_CARD_PREFIX = "WELLS FARGO ACTIVE CASH VISA(R) CARD"
WF_CARD_ALIAS = "WFACV"

DATE_FORMATS = (
    "%m/%d/%Y",
    "%m/%d/%y",
    "%Y-%m-%d",
    "%m-%d-%Y",
    "%m-%d-%y",
)


# -----------------------------
# Helpers
# -----------------------------
def normalize_spaces(text: str) -> str:
    """Trim and collapse internal whitespace to a single space."""
    return " ".join((text or "").split()).strip()


def normalize_payment_method(value: str) -> str:
    """
    Normalize Wells Fargo Active Cash Visa card names:
    'WELLS FARGO ACTIVE CASH VISA(R) CARD ...4321' -> 'WFACV...4321'
    """
    value = (value or "").strip()
    if value.upper().startswith(WF_CARD_PREFIX):
        suffix = value[len(WF_CARD_PREFIX):].strip()  # keeps "...4321"
        return f"{WF_CARD_ALIAS}{suffix}"
    return value


def parse_date(value: str):
    """Parse various common date formats. Returns datetime or None."""
    s = ("" if value is None else str(value)).strip()
    if not s:
        return None

    # remove time part if present (e.g., "2026-01-05 00:00:00")
    s = s.split()[0]
    # handle ISO "2026-01-05T00:00:00"
    if "T" in s:
        s = s.split("T")[0]

    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


# -----------------------------
# Core processing
# -----------------------------
def load_rows(csv_path: Path):
    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        return reader.fieldnames, list(reader)


def clean_rows(rows):
    removed = 0
    cleaned = []

    for row in rows:
        # Normalize Description spacing
        row["Description"] = normalize_spaces(row.get("Description"))

        # Remove ONLINE TRANSFER REF... rows (case-insensitive)
        desc_key = (row.get("Description") or "").upper()
        if desc_key.startswith(REMOVE_DESC_PREFIX.upper()):
            removed += 1
            continue

        # Normalize Payment Method
        row["Payment Method"] = normalize_payment_method(row.get("Payment Method"))

        cleaned.append(row)

    return cleaned, removed


def sort_rows(rows):
    # Sort by Description A→Z then by Date (unparsed dates go last within group)
    rows.sort(
        key=lambda r: (
            (r.get("Description") or "").upper(),
            parse_date(r.get("Date")) or datetime.max,
        )
    )
    return rows


def write_xlsx(headers, rows, xlsx_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Cleaned & Sorted"

    ws.append(headers)
    for r in rows:
        ws.append([r.get(h, "") for h in headers])

    wb.save(xlsx_path)


def main():
    base_dir = Path(__file__).parent
    csv_path = base_dir / INPUT_CSV
    xlsx_path = base_dir / OUTPUT_XLSX

    if not csv_path.exists():
        raise FileNotFoundError(f"Input CSV not found: {csv_path}")

    headers, rows = load_rows(csv_path)
    cleaned, removed = clean_rows(rows)
    sorted_rows = sort_rows(cleaned)
    write_xlsx(headers, sorted_rows, xlsx_path)

    # Debug: how many dates parsed successfully
    parsed_ok = sum(1 for r in sorted_rows if parse_date(r.get("Date")) is not None)

    print("✅ Done")
    print(f"🧹 Removed rows (prefix '{REMOVE_DESC_PREFIX}'): {removed}")
    print(f"🔤 Sorted: Description (A→Z) then Date (oldest→newest)")
    print(f"📅 Date parsed successfully: {parsed_ok}/{len(sorted_rows)}")
    print(f"📁 Output: {xlsx_path.name}")


if __name__ == "__main__":
    main()
import csv
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook

def normalize_description(value):
    """
    Fix inconsistent spacing in Description:
    - strip leading/trailing spaces
    - collapse multiple spaces into one
    """
    if not value:
        return value
    return " ".join(value.split())


def clean_payment_method(value):
    if not value:
        return value

    value = value.strip()
    prefix = "WELLS FARGO ACTIVE CASH VISA(R) CARD"

    if value.upper().startswith(prefix):
        suffix = value[len(prefix):].strip()
        return f"WFACV{suffix}"

    return value


def parse_date(date_str):
    if not date_str:
        return None

    s = str(date_str).strip()
    s = s.split()[0]  # remove time if present
    if "T" in s:
        s = s.split("T")[0]

    formats = [
        "%m/%d/%Y",
        "%m/%d/%y",
        "%Y-%m-%d",
        "%m-%d-%Y",
        "%m-%d-%y",
    ]

    for fmt in formats:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue

    return None


def sort_clean_and_normalize(csv_filename):
    base_dir = Path(__file__).parent
    input_csv = base_dir / csv_filename
    output_xlsx = base_dir / "expenses_sorted_A_Z_by_description_then_date.xlsx"

    if not input_csv.exists():
        print(f"❌ File not found: {input_csv}")
        return

    with open(input_csv, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        headers = reader.fieldnames

    cleaned_rows = []
    removed_count = 0

    for row in rows:
        # 🔧 Normalize Description spacing
        raw_desc = row.get("Description")
        row["Description"] = normalize_description(raw_desc)

        desc_key = (row["Description"] or "").upper()

        # ❌ Remove ONLINE TRANSFER REF rows
        if desc_key.startswith("ONLINE TRANSFER REF"):
            removed_count += 1
            continue

        # 🔁 Normalize Payment Method
        row["Payment Method"] = clean_payment_method(row.get("Payment Method"))

        cleaned_rows.append(row)

    # 🔤 Sort by Description A→Z, then 📅 by Date
    cleaned_rows.sort(
        key=lambda x: (
            (x.get("Description") or "").upper(),
            parse_date(x.get("Date")) or datetime.max
        )
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Sorted & Cleaned"

    ws.append(headers)
    for row in cleaned_rows:
        ws.append([row.get(h, "") for h in headers])

    wb.save(output_xlsx)

    print("✅ Processing complete")
    print(f"🧹 Rows removed (ONLINE TRANSFER REF): {removed_count}")
    print("🔤 Description spacing normalized")
    print("📅 Sorted by Description → Date")
    print(f"📁 Output file: {output_xlsx.name}")


if __name__ == "__main__":
    sort_clean_and_normalize("expenses.csv")
import csv
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook

def clean_payment_method(value):
    if not value:
        return value

    value = value.strip()
    prefix = "WELLS FARGO ACTIVE CASH VISA(R) CARD"

    if value.upper().startswith(prefix):
        suffix = value[len(prefix):].strip()  # keeps "...4321"
        return f"WFACV{suffix}"

    return value


def parse_date(date_str):
    """
    Robust date parser for common bank/Excel exports.
    Handles:
      - 01/05/2026
      - 1/5/26
      - 2026-01-05
      - 2026-01-05 00:00:00
      - 01/05/2026 12:34:56
      - ISO like 2026-01-05T00:00:00
    """
    if not date_str:
        return None

    s = str(date_str).strip()

    # If there's a time component, keep only date part (first token)
    # Works for "2026-01-05 00:0
import csv
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook

def clean_payment_method(value):
    """
    Normalize Wells Fargo Active Cash Visa card names
    """
    if not value:
        return value

    value = value.strip()
    prefix = "WELLS FARGO ACTIVE CASH VISA(R) CARD"

    if value.upper().startswith(prefix):
        suffix = value[len(prefix):].strip()
        return f"WFACV{suffix}"

    return value


def parse_date(date_str):
    """
    Safely parse date for sorting.
    Adjust formats here if needed.
    """
    if not date_str:
        return datetime.min

    for fmt in ("%m/%d/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(date_str.strip(), fmt)
        except ValueError:
            pass

    return datetime.min


def sort_clean_and_normalize(csv_filename):
    base_dir = Path(__file__).parent
    input_csv = base_dir / csv_filename
    output_xlsx = base_dir / "expenses_sorted_A_Z_by_description_then_date.xlsx"

    if not input_csv.exists():
        print(f"❌ File not found: {input_csv}")
        return

    with open(input_csv, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        headers = reader.fieldnames

    original_count = len(rows)
    cleaned_rows = []
    removed_count = 0

    for row in rows:
        description = (row.get("Description") or "").strip().upper()

        # ❌ Remove ONLINE TRANSFER REF
        if description.startswith("ONLINE TRANSFER REF"):
            removed_count += 1
            continue

        # 🔁 Normalize Payment Method
        row["Payment Method"] = clean_payment_method(row.get("Payment Method"))

        cleaned_rows.append(row)

    # 🔤 Sort by Description A→Z, then 📅 by Date
    cleaned_rows.sort(
        key=lambda x: (
            (x.get("Description") or "").strip().upper(),
            parse_date(x.get("Date"))
        )
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Sorted by Description & Date"

    ws.append(headers)
    for row in cleaned_rows:
        ws.append([row[h] for h in headers])

    wb.save(output_xlsx)

    print("✅ Processing complete")
    print(f"🧹 Rows removed (ONLINE TRANSFER REF): {removed_count}")
    print("🔤 Sorted by Description (A→Z)")
    print("📅 Then sorted by Date (oldest → newest)")
    print(f"📁 Output file: {output_xlsx.name}")


if __name__ == "__main__":
    sort_clean_and_normalize("expenses.csv")

import csv
from pathlib import Path
from openpyxl import Workbook

def sort_and_remove_online_transfers(csv_filename):
    base_dir = Path(__file__).parent
    input_csv = base_dir / csv_filename
    output_xlsx = base_dir / "expenses_sorted_A_Z_no_online_transfer.xlsx"

    if not input_csv.exists():
        print(f"❌ File not found: {input_csv}")
        return

    with open(input_csv, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        headers = reader.fieldnames

    original_count = len(rows)

    # 🧹 REMOVE rows starting with "ONLINE TRANSFER REF"
    filtered_rows = [
        row for row in rows
        if not (row["Description"] or "").strip().upper().startswith("ONLINE TRANSFER REF")
    ]

    removed_count = original_count - len(filtered_rows)

    # 🔤 SORT A → Z by Description
    filtered_rows.sort(
        key=lambda x: (x["Description"] or "").strip().upper()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Sorted A-Z (Cleaned)"

    ws.append(headers)
    for row in filtered_rows:
        ws.append([row[h] for h in headers])

    wb.save(output_xlsx)

    print("✅ Cleanup + Sort completed")
    print(f"🧹 Rows removed: {removed_count}")
    print(f"📁 Output file: {output_xlsx.name}")

if __name__ == "__main__":
    sort_and_remove_online_transfers("expenses.csv")
import csv
from pathlib import Path
from openpyxl import Workbook

def sort_csv_by_description_to_excel(csv_filename):
    base_dir = Path(__file__).parent
    input_csv = base_dir / csv_filename
    output_xlsx = base_dir / "expenses_sorted_A_Z.xlsx"

    if not input_csv.exists():
        print(f"❌ File not found: {input_csv}")
        return

    with open(input_csv, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        headers = reader.fieldnames

    # ✅ TRUE A → Z SORT
    rows.sort(
        key=lambda x: (x["Description"] or "").strip().upper()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Sorted A-Z"

    ws.append(headers)
    for row in rows:
        ws.append([row[h] for h in headers])

    wb.save(output_xlsx)
    print("✅ Sorted A → Z by Description")
    print(f"📁 Output: {output_xlsx.name}")

if __name__ == "__main__":
    sort_csv_by_description_to_excel("expenses.csv")
import csv
from pathlib import Path
from openpyxl import Workbook

def sort_csv_by_description_to_excel(csv_filename):
    base_dir = Path(__file__).parent
    input_csv = base_dir / csv_filename
    output_xlsx = base_dir / input_csv.with_suffix(".xlsx").name

    if not input_csv.exists():
        print(f"❌ File not found: {input_csv}")
        return

    with open(input_csv, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        rows = list(reader)
        headers = reader.fieldnames

    # Sort by Description (case-insensitive)
    rows.sort(key=lambda x: (x["Description"] or "").lower())

    wb = Workbook()
    ws = wb.active
    ws.title = "Sorted Data"

    ws.append(headers)
    for row in rows:
        ws.append([row[h] for h in headers])

    wb.save(output_xlsx)
    print(f"✅ Excel file created in same folder: {output_xlsx.name}")

if __name__ == "__main__":
    # Change this if your CSV name is different
    sort_csv_by_description_to_excel("expenses.csv")

